"""
Unit Tests for TARA Excel Generator
====================================

This test module verifies that the Excel generator produces reports matching
the format of MY25 EV平台中控主机_TARA分析报告.xlsx.

The test constructs input data for the non-auto-calculated columns and verifies:
1. Sheet structure matches the sample
2. Column headers are correct
3. Formulas are correctly placed in auto-calculated columns
4. Data rows contain expected values

Auto-calculated columns in 5-TARA分析结果 (these are formulas, not input data):
- M(13): Attack Vector Value
- O(15): Attack Complexity Value  
- Q(17): Privileges Required Value
- S(19): User Interaction Value
- T(20): Attack Feasibility Calculation
- U(21): Attack Feasibility Level
- W(23): Safety Notes
- X(24): Safety Value
- Z(26): Financial Notes
- AA(27): Financial Value
- AC(29): Operational Notes
- AD(30): Operational Value
- AF(32): Privacy Notes
- AG(33): Privacy Value
- AH(34): Impact Calculation
- AI(35): Impact Level
- AJ(36): Risk Level
- AK(37): Risk Treatment Decision
- AL(38): Security Goal
- AN(40): WP29 Control Mapping
"""

import asyncio
import os
import sys
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.generators.excel_generator import ExcelGenerator


# Sample input data matching MY25 EV平台中控主机_TARA分析报告.xlsx
# Note: Only input columns are included, auto-calculated columns are handled by formulas

SAMPLE_PROJECT = {
    "name": "MY25 EV平台中控主机",
    "platform_name": "DiLink150中控主机平台",
    "description": "车载信息娱乐系统(In-Vehicle Infotainment, IVI)是一种集成多媒体娱乐、导航、车辆信息显示、通信和车辆控制功能于一体的车载电子系统。",
    "doc_number": "IPC0011_JF_A30-44003",
    "version": "1.0",
    "author_date": "2025.11",
    "review_date": "2025.0",
}

SAMPLE_ASSETS = [
    {
        "id": "P001",
        "name": "SOC",
        "category": "内部实体",
        "description": "主处理器芯片，运行Android Automotive OS",
        "security_attrs": {
            "authenticity": True,
            "integrity": False,
            "non_repudiation": False,
            "confidentiality": False,
            "availability": True,
            "authorization": False,
        }
    },
    {
        "id": "P011",
        "name": "媒体源",
        "category": "固件及应用",
        "description": "对媒体源的分析包括其本身的代码和相关的应用服务，负责接收视频及人脸信息，发送至存储单元",
        "security_attrs": {
            "authenticity": True,
            "integrity": True,
            "non_repudiation": False,
            "confidentiality": False,
            "availability": True,
            "authorization": False,
        }
    },
    {
        "id": "E014",
        "name": "TBox",
        "category": "外部实体",
        "description": "与DiLink主机连接的外部设备",
        "security_attrs": {
            "authenticity": True,
        }
    },
    {
        "id": "E009",
        "name": "外置功放",
        "category": "外部实体",
        "description": "与DiLink主机连接的外部设备",
        "security_attrs": {
            "authenticity": True,
        }
    },
    {
        "id": "E004",
        "name": "以太网接口",
        "category": "外部实体",
        "description": "用于以太网通讯、调试、数据传输的接口",
        "security_attrs": {
            "authorization": True,
        }
    },
    {
        "id": "E005",
        "name": "UART接口",
        "category": "外部实体",
        "description": "用于调试、通讯的接口",
        "security_attrs": {
            "authorization": True,
        }
    },
    {
        "id": "S005",
        "name": "OTA升级包",
        "category": "数据存储",
        "description": "存储在DDR中",
        "security_attrs": {
            "authenticity": True,
            "integrity": True,
            "confidentiality": True,
            "authorization": True,
        }
    },
    {
        "id": "D006",
        "name": "音频信息",
        "category": "数据流",
        "description": "MCU<--->音频输入输出模块",
        "security_attrs": {
            "authenticity": True,
            "integrity": True,
            "confidentiality": True,
        }
    },
    {
        "id": "A-009",
        "name": "系统固件",
        "category": "固件及应用",
        "description": "Bootloader、TF-A、TEE OS等",
        "security_attrs": {
            "integrity": True,
            "non_repudiation": True,
            "availability": True,
        }
    },
]

# TARA analysis results - INPUT DATA ONLY (no auto-calculated columns)
SAMPLE_THREATS = [
    {
        "asset_id_str": "P001",
        "asset_name": "车载多媒体",
        "category_sub1": "系统实体",
        "category_sub2": "N/A",
        "category_sub3": "SOC",
        "category": "内部实体",
        "security_attribute": '"Authenticity | 真实性"',
        "stride_model": "S欺骗",
        "threat_scenario": "黑客仿冒、篡改SOC模块，导致零部件内部所有与SOC关联的模块无法接收真实的信息正常完成功能，甚至会使攻击者成功通过重要身份认证，造成安全隐患",
        "attack_path": "1.攻击者锁定待攻击车辆，通过物理拆解取得芯片，并进行调查，了解芯片的物理结构、通讯协议等相关信息。\n2.通过观察寻找芯片可读丝印，或进行X光扫描，寻找可进",
        "wp29_mapping": "4.1\n5.1\n6.1\n6.2\n6.3\n",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "1.应确保操作系统启动时应使用可信机制，在验证操作系统签名并判定通过后，再从可信存储区域加载操作系统，避免加载被算改的操作系统",
    },
    {
        "asset_id_str": "",
        "asset_name": "",
        "category_sub1": "",
        "category_sub2": "",
        "category_sub3": "SOC",
        "category": "",
        "security_attribute": '"Availability | 可用性"',
        "stride_model": "D拒绝服务",
        "threat_scenario": "黑客通过超载SOC模块的处理能力或发送大量无效请求，造成SOC模块过载，导致车辆系统崩溃或无响应，导致无法向零部件模块接收或传递数据，功能无法正常通过系统使用。",
        "attack_path": "1.攻击者通过物理手段（如高温、高压、电磁辐射等）对MCU进行破坏，\n2.损坏硬件组件或改变电子状态，导致SOC无法继续工作。",
        "wp29_mapping": "24.1\n8.1",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "可忽略不计的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "1、操作系统应裁剪掉不需要的组件与原生apk应用，如不对量产开放的开发工具等，以防止不必要的资源占用并防止攻击者利用开发工具恶意破坏系统",
    },
    {
        "asset_id_str": "P011",
        "asset_name": "车载多媒体",
        "category_sub1": "SoC",
        "category_sub2": "Flash",
        "category_sub3": "媒体源",
        "category": "固件及应用",
        "security_attribute": '"Authenticity | 真实性"',
        "stride_model": "S欺骗",
        "threat_scenario": "多媒体源数据被仿冒，导致多媒体真实性被破坏，导致多媒体模块传递的错误的视频信息，产生误导影响。",
        "attack_path": "1.攻击者通过USB接口攻击多媒体主机，获得多媒体主机权限，进而攻击其他车载部件\n2.使用伪造的固件及应用进行替换\n3.导致功能执行异常",
        "wp29_mapping": "4.1\n4.2",
        "attack_vector": "本地",
        "attack_complexity": "高",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "",
        "asset_name": "",
        "category_sub1": "",
        "category_sub2": "",
        "category_sub3": "媒体源",
        "category": "固件及应用",
        "security_attribute": '"Integrity | 完整性"',
        "stride_model": "T篡改",
        "threat_scenario": "多媒体源被篡改，攻击者可以修改视频内容，导致对车辆安全和操控产生负面影响。",
        "attack_path": "1.攻击者通过硬件调试接口如JTAG、UART接口攻击车载部件\n2.车载部件固件逻辑被篡改\n3.执行器错误执行相关指令，导致功能执行异常",
        "wp29_mapping": "4.1\n5.1\n6.1\n6.2\n6.3",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "需要",
        "safety_impact": "可忽略不计的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "基于媒体文件的格式定义，对文件格式进行符合性校验，只允许读写指定格式的文件或安装执行指定签名的应用软件",
    },
    {
        "asset_id_str": "",
        "asset_name": "",
        "category_sub1": "",
        "category_sub2": "",
        "category_sub3": "媒体源",
        "category": "固件及应用",
        "security_attribute": '"Availability | 可用性"',
        "stride_model": "D拒绝服务",
        "threat_scenario": "多媒体源收到攻击者拒绝服务攻击或攻击者通过发送大量无效请求或恶意指令，使多媒体源过载或崩溃，导致视频信息无法正常传输或显示，导致驾驶员对周围环境的感知能力，增加",
        "attack_path": "1.攻击者发送大量无效、恶意或异常的请求、命令或数据包，使零部件并发执行大量任务，占用CPU、内存、网络带宽等各类资源资源。\n2.功能模块在处理过多恶意请求",
        "wp29_mapping": "24.1\n8.1",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "高",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "E014\n",
        "asset_name": "车载多媒体",
        "category_sub1": "系统实体",
        "category_sub2": "N/A",
        "category_sub3": "OBD接口",
        "category": "外部实体",
        "security_attribute": '"Elevation of Privilege | 权限提升"',
        "stride_model": "E权限提升",
        "threat_scenario": "攻击者通过非授权的诊断仪等设备连接OBD端口，通过恶意指令执行操作，破坏系统安全。",
        "attack_path": "1、攻击者通过相关非授权设备恶意连接OBD接口，破坏系统安全。",
        "wp29_mapping": "7.1\n9.1",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "通过扫二维码等方式对OBD接口实行访问控制，不允许非受控的连接。",
    },
    {
        "asset_id_str": "E009",
        "asset_name": "车载多媒体",
        "category_sub1": "系统实体",
        "category_sub2": "N/A",
        "category_sub3": "外置功放",
        "category": "外部实体",
        "security_attribute": '"Authenticity | 真实性"',
        "stride_model": "S欺骗",
        "threat_scenario": "与多媒体主机连接的外置功放被替换，导致功能调用的时候发生异常",
        "attack_path": "1、攻击者通过拆卸外部设备，分析设备。\n2、使用其他零部件替换外部实体设备。\n3、导致功能执行异常",
        "wp29_mapping": "27.1\n28.2\n30.1\n32.1",
        "attack_vector": "物理",
        "attack_complexity": "低",
        "privileges_required": "无",
        "user_interaction": "不需要",
        "safety_impact": "可忽略不计的",
        "financial_impact": "可忽略不计的",
        "operational_impact": "中等的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "E004",
        "asset_name": "车载多媒体",
        "category_sub1": "系统实体",
        "category_sub2": "N/A",
        "category_sub3": "UART接口",
        "category": "外部实体",
        "security_attribute": '"Elevation of Privilege | 权限提升"',
        "stride_model": "E权限提升",
        "threat_scenario": "攻击者冒充外部合法设备连接UART接口进行各种攻击，可能导致零部件无法使用。",
        "attack_path": "攻击者通过相关非授权设备连接UART接口，通过指令实行访问、攻击等操作。",
        "wp29_mapping": "7.1\n9.1",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "中等的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "E005",
        "asset_name": "车载多媒体",
        "category_sub1": "系统实体",
        "category_sub2": "N/A",
        "category_sub3": "CAN/CANFD接口",
        "category": "外部实体",
        "security_attribute": '"Elevation of Privilege | 权限提升"',
        "stride_model": "E权限提升",
        "threat_scenario": "攻击者冒充外部合法设备连接CAN/CANFD接口进行各种攻击，可能导致零部件无法使用。",
        "attack_path": "攻击者通过诊断仪等非授权设备连接CAN接口，通过指令实行访问、攻击等操作。",
        "wp29_mapping": "7.1\n9.1",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "严重的",
        "financial_impact": "中等的",
        "operational_impact": "中等的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "S005",
        "asset_name": "车载多媒体",
        "category_sub1": "片外储存",
        "category_sub2": "DDR",
        "category_sub3": "OTA升级包",
        "category": "数据存储",
        "security_attribute": '"Authenticity | 真实性"',
        "stride_model": "S欺骗",
        "threat_scenario": "储存在片外储存的OTA升级包被替换，导致功能调用的时候发生异常",
        "attack_path": "1.攻击者通过OBD接口攻击车载部件，进而非法访问存储在车载部件的软件升级包\n2.使用伪造的故障和保养信息数据替换原来的故障和保养信息\n3.故障和保养信",
        "wp29_mapping": "4.1\n4.2\n12.1",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "中等的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "",
        "asset_name": "",
        "category_sub1": "",
        "category_sub2": "",
        "category_sub3": "OTA升级包",
        "category": "数据存储",
        "security_attribute": '"Integrity | 完整性"',
        "stride_model": "T篡改",
        "threat_scenario": "攻击者恶意篡改系统DDR中存储的OTA升级包数据，导致OTA升级失败，无法实现其正常功能",
        "attack_path": "1. 攻击者利用专用的转接设备连接到车辆的外部接口（如USB端口、OBD端口或JTAG口）。\n2. 攻击者提取OTA升级包进行篡改并将篡改后的OTA升级包刷",
        "wp29_mapping": "26.1\n26.2\n26.3",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "",
        "asset_name": "",
        "category_sub1": "",
        "category_sub2": "",
        "category_sub3": "OTA升级包",
        "category": "数据存储",
        "security_attribute": '"Confidentiality | 机密性"',
        "stride_model": "I信息泄露",
        "threat_scenario": "1.攻击者通过USB接口攻击多媒体主机，获得多媒体主机权限，进而攻击其他车载部件\n2.读取固件应用的敏感信息\n3.导致车辆敏感信息泄露",
        "attack_path": "1.攻击者通过USB接口攻击多媒体主机\n2.访问本地存储的OTA升级包，篡改本地OTA升级包\n3.导致OTA升级包泄露",
        "wp29_mapping": "7.1\n3.5\n31.1",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "可忽略不计的",
        "financial_impact": "重大的",
        "operational_impact": "可忽略不计的",
        "privacy_impact": "重大的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "",
        "asset_name": "",
        "category_sub1": "",
        "category_sub2": "",
        "category_sub3": "OTA升级包",
        "category": "数据存储",
        "security_attribute": '"Elevation of Privilege | 权限提升"',
        "stride_model": "E权限提升",
        "threat_scenario": "攻击者越权访问存储器，使得OTA升级包存储器机密性被破坏，攻击者可根据获取的升级包进行篡改、注入等一系列活动。",
        "attack_path": "1. 攻击者锁定待攻击车辆，通过物理拆解取得芯片，并进行调查，了解芯片的物理结构、通讯协议等相关信息。\n2. 通过观察寻找芯片可读丝印，或进行X光扫描，寻找",
        "wp29_mapping": "13.1\n8.1",
        "attack_vector": "本地",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "D006",
        "asset_name": "车载多媒体",
        "category_sub1": "通讯接口",
        "category_sub2": "无线传输",
        "category_sub3": "BT模块控制指令",
        "category": "数据流",
        "security_attribute": '"Authenticity | 真实性"',
        "stride_model": "S欺骗",
        "threat_scenario": "攻击者恶意仿冒蓝牙模块控制指令，导致系统执行虚假的指令，破坏车辆安全。",
        "attack_path": "1. 攻击者锁定目标车辆，找到负责通信的零部件位置，拨开线束并拆解下零部件。\n2. 利用已经找到的接口或调试点，通过飞线，使用定制转接器连接到自己的电脑设备",
        "wp29_mapping": "24.1\n8.1",
        "attack_vector": "邻居",
        "attack_complexity": "高",
        "privileges_required": "低",
        "user_interaction": "需要",
        "safety_impact": "可忽略不计的",
        "financial_impact": "中等的",
        "operational_impact": "中等的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "",
        "asset_name": "",
        "category_sub1": "",
        "category_sub2": "",
        "category_sub3": "BT模块控制指令",
        "category": "数据流",
        "security_attribute": '"Non-Repudiation | 不可抵赖性"',
        "stride_model": "R抵赖",
        "threat_scenario": "攻击者攻击多媒体系统，发送虚假的控制指令，系统接收到错误信息无法溯源，进而影响蓝牙模块使用安全。",
        "attack_path": "1. 攻击者锁定目标车辆，找到负责通信的零部件位置，拨开线束并拆解下零部件。\n2. 利用已经找到的接口或调试点，通过飞线，使用定制转接器连接到自己的电脑设备",
        "wp29_mapping": "4.1\n4.2",
        "attack_vector": "网络",
        "attack_complexity": "低",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "可忽略不计的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "中等的",
        "security_requirement": "",
    },
    {
        "asset_id_str": "",
        "asset_name": "",
        "category_sub1": "",
        "category_sub2": "",
        "category_sub3": "BT模块控制指令",
        "category": "数据流",
        "security_attribute": '"Integrity | 完整性"',
        "stride_model": "T篡改",
        "threat_scenario": "攻击者对BT模块控制指令进行篡改，导致相关零部件收到了错误的指令，不能按预期工作。",
        "attack_path": "1. 攻击者锁定目标车辆，找到负责通信的零部件位置，拨开线束并拆解下零部件。\n2. 利用已经找到的接口或调试点，通过飞线，使用定制转接器连接到自己的电脑设备",
        "wp29_mapping": "5.2\n6.2\n11.3",
        "attack_vector": "邻居",
        "attack_complexity": "高",
        "privileges_required": "低",
        "user_interaction": "不需要",
        "safety_impact": "中等的",
        "financial_impact": "中等的",
        "operational_impact": "重大的",
        "privacy_impact": "可忽略不计的",
        "security_requirement": "",
    },
]

SAMPLE_MEASURES = []  # No measures in the sample data for this test


class TestExcelGeneratorTARA:
    """Test suite for TARA Excel Generator."""

    @pytest.fixture
    def generator(self):
        """Create an ExcelGenerator instance."""
        return ExcelGenerator()

    @pytest.fixture
    def sample_data(self):
        """Create sample data matching the sample Excel."""
        return {
            "content": {
                "project": SAMPLE_PROJECT,
                "assets": SAMPLE_ASSETS,
                "threats": SAMPLE_THREATS,
                "control_measures": SAMPLE_MEASURES,
            }
        }

    @pytest.mark.asyncio
    async def test_generate_creates_all_sheets(self, generator, sample_data):
        """Test that all required sheets are created."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer)
        
        expected_sheets = [
            "0. 封面 Front Cover",
            "1-相关定义",
            "2-资产列表",
            "3-数据流图",
            "4-攻击树图",
            "5-TARA分析结果",
        ]
        
        assert wb.sheetnames == expected_sheets

    @pytest.mark.asyncio
    async def test_tara_results_has_correct_columns(self, generator, sample_data):
        """Test that 5-TARA分析结果 has 40 columns."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer)
        ws = wb["5-TARA分析结果"]
        
        # Should have 40 columns
        assert ws.max_column == 40

    @pytest.mark.asyncio
    async def test_tara_results_has_formulas(self, generator, sample_data):
        """Test that auto-calculated columns contain formulas."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer, data_only=False)
        ws = wb["5-TARA分析结果"]
        
        # Check row 6 (first data row) for formulas
        formula_columns = {
            13: "M",  # Attack Vector Value
            15: "O",  # Attack Complexity Value
            17: "Q",  # Privileges Required Value
            19: "S",  # User Interaction Value
            20: "T",  # Attack Feasibility Calculation
            21: "U",  # Attack Feasibility Level
            23: "W",  # Safety Notes
            24: "X",  # Safety Value
            26: "Z",  # Financial Notes
            27: "AA", # Financial Value
            29: "AC", # Operational Notes
            30: "AD", # Operational Value
            32: "AF", # Privacy Notes
            33: "AG", # Privacy Value
            34: "AH", # Impact Calculation
            35: "AI", # Impact Level
            36: "AJ", # Risk Level
            37: "AK", # Risk Treatment Decision
            38: "AL", # Security Goal
            40: "AN", # WP29 Control Mapping
        }
        
        for col, col_name in formula_columns.items():
            cell = ws.cell(row=6, column=col)
            assert isinstance(cell.value, str), f"Column {col_name}({col}) should have a formula"
            assert cell.value.startswith("="), f"Column {col_name}({col}) formula should start with ="

    @pytest.mark.asyncio
    async def test_tara_results_input_columns_have_values(self, generator, sample_data):
        """Test that input columns contain the expected values."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer)
        ws = wb["5-TARA分析结果"]
        
        # Check first data row (row 6)
        assert ws.cell(row=6, column=1).value == "P001"  # Asset ID
        assert ws.cell(row=6, column=8).value == "S欺骗"  # STRIDE Model
        assert ws.cell(row=6, column=12).value == "本地"  # Attack Vector
        assert ws.cell(row=6, column=14).value == "低"    # Attack Complexity
        assert ws.cell(row=6, column=16).value == "低"    # Privileges Required
        assert ws.cell(row=6, column=18).value == "不需要" # User Interaction
        assert ws.cell(row=6, column=22).value == "中等的" # Safety Impact
        assert ws.cell(row=6, column=25).value == "中等的" # Financial Impact
        assert ws.cell(row=6, column=28).value == "重大的" # Operational Impact
        assert ws.cell(row=6, column=31).value == "可忽略不计的" # Privacy Impact

    @pytest.mark.asyncio
    async def test_tara_results_row_count(self, generator, sample_data):
        """Test that the correct number of data rows are generated."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer)
        ws = wb["5-TARA分析结果"]
        
        # Header rows: 1 (title) + 2 (empty) + 3 (group headers) + 4-5 (sub-headers)
        # Data starts at row 6
        expected_data_rows = len(SAMPLE_THREATS)
        # max_row should be 5 (headers) + number of threats
        assert ws.max_row == 5 + expected_data_rows

    @pytest.mark.asyncio
    async def test_asset_list_sheet(self, generator, sample_data):
        """Test that 2-资产列表 sheet is generated correctly."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer)
        ws = wb["2-资产列表"]
        
        # Check title
        assert "资产列表" in str(ws.cell(row=1, column=1).value)
        
        # Check headers (row 4)
        assert "资产ID" in str(ws.cell(row=4, column=1).value)
        assert "资产名称" in str(ws.cell(row=4, column=2).value)
        
        # Check first data row
        assert ws.cell(row=5, column=1).value == "P001"
        assert ws.cell(row=5, column=2).value == "SOC"

    @pytest.mark.asyncio
    async def test_formula_attack_vector_value(self, generator, sample_data):
        """Test the Attack Vector Value formula."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer, data_only=False)
        ws = wb["5-TARA分析结果"]
        
        # Check formula in M6
        formula = ws.cell(row=6, column=13).value
        assert '=IF(L6="网络",0.85' in formula
        assert '=IF(L6="邻居",0.62' in formula or 'L6="邻居",0.62' in formula
        assert '=IF(L6="本地",0.55' in formula or 'L6="本地",0.55' in formula
        assert '=IF(L6="物理",0.2' in formula or 'L6="物理",0.2' in formula

    @pytest.mark.asyncio
    async def test_formula_attack_feasibility(self, generator, sample_data):
        """Test the Attack Feasibility Calculation formula."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer, data_only=False)
        ws = wb["5-TARA分析结果"]
        
        # Check formula in T6
        formula = ws.cell(row=6, column=20).value
        assert formula == "=8.22*M6*O6*Q6*S6"

    @pytest.mark.asyncio
    async def test_formula_impact_calculation(self, generator, sample_data):
        """Test the Impact Calculation formula."""
        buffer = await generator.generate(sample_data)
        wb = load_workbook(buffer, data_only=False)
        ws = wb["5-TARA分析结果"]
        
        # Check formula in AH6
        formula = ws.cell(row=6, column=34).value
        assert formula == "=SUM(X6+AA6+AD6+AG6)"

    @pytest.mark.asyncio
    async def test_generate_and_save_file(self, generator, sample_data, tmp_path):
        """Test generating and saving an Excel file."""
        buffer = await generator.generate(sample_data)
        
        # Save to file
        output_path = tmp_path / "test_tara_report.xlsx"
        with open(output_path, "wb") as f:
            f.write(buffer.getvalue())
        
        assert output_path.exists()
        assert output_path.stat().st_size > 0
        
        # Reload and verify
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) == 6
