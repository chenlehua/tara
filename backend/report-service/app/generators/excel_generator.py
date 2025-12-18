"""Excel report generator - ISO 21434 TARA Format.

This generator creates Excel reports that match the MY25 EV平台中控主机_TARA分析报告.xlsx format,
including proper formulas for auto-calculated columns.
"""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from app.common.utils import get_logger
    logger = get_logger(__name__)
except ImportError:
    import logging
    logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel reports in ISO 21434 TARA format.
    
    This generator matches the format of MY25 EV平台中控主机_TARA分析报告.xlsx,
    including proper Excel formulas for auto-calculated columns.
    """

    def __init__(self):
        """Initialize styles matching the sample Excel format."""
        # Title style (Row 1)
        self.title_font = Font(bold=True, size=14, color="2F5496")
        self.title_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Header group style (Row 3)
        self.header_font = Font(bold=True, size=10, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        
        # Sub-header style (Row 4)
        self.subheader_font = Font(bold=True, size=10)
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Detail header style (Row 5)
        self.detail_font = Font(bold=True, size=9)
        self.detail_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
        
        # Data font
        self.data_font = Font(size=9)
        
        # Border
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Alignments
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Risk level colors (for conditional formatting)
        self.risk_fills = {
            'Critical': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            'High': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'Medium': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            'Low': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            'QM': PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
        }

    def _apply_cell_style(self, cell, font=None, fill=None, alignment=None, border=True):
        """Apply styles to a cell."""
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = self.border

    def _merge_and_style(self, ws, start_row, start_col, end_row, end_col, value, font=None, fill=None, alignment=None):
        """Merge cells and apply style."""
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col, value=value)
        self._apply_cell_style(cell, font, fill, alignment or self.center_align)
        
        # Apply border to all merged cells
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = self.border

    async def generate(
        self,
        data: dict,
        template: str = "iso21434",
        sections: list[str] = None,
    ) -> io.BytesIO:
        """Generate Excel report in TARA format.
        
        Args:
            data: Dictionary containing project, assets, threats, and control_measures
            template: Template type (default: iso21434)
            sections: Optional list of sections to include
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        
        # Get content
        content = data.get("content", data)
        project = content.get("project", data.get("project", {}))
        assets = content.get("assets", [])
        threats = content.get("threats", [])
        measures = content.get("control_measures", [])

        # Create sheets in order
        self._create_cover_sheet(wb, project)
        self._create_definitions_sheet(wb, project)
        self._create_asset_list_sheet(wb, assets, project)
        self._create_data_flow_sheet(wb, assets, project)
        self._create_attack_tree_sheet(wb, threats, project)
        self._create_tara_results_sheet(wb, assets, threats, measures, project)

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _create_cover_sheet(self, wb: Workbook, project: dict):
        """Create 0. 封面 Front Cover sheet."""
        ws = wb.create_sheet("0. 封面 Front Cover", 0)
        
        project_name = project.get("name", "TARA分析项目")
        platform_name = project.get("platform_name", project_name)
        
        # Data classification
        ws.cell(row=4, column=6, value="数据等级：秘密\nData level: Confidential")
        ws.cell(row=5, column=6, value=f"编号：{project.get('doc_number', 'IPC0011_JF_A30-44003')}\nNumber: {project.get('doc_number', 'IPC0011_JF_A30-44003')}")
        ws.cell(row=6, column=6, value=f"版本：{project.get('version', '1.0')}\nVersion：{project.get('version', '1.0')}")
        
        # Title
        title_cell = ws.cell(row=7, column=1, value="威胁分析和风险评估报告\nThreat Analysis And Risk Assessment Report")
        title_cell.font = Font(bold=True, size=24)
        title_cell.alignment = self.center_align
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=7)
        
        # Platform subtitle
        subtitle_cell = ws.cell(row=8, column=5, value=f"——{platform_name}")
        subtitle_cell.font = Font(size=16)
        
        # Author/Date info
        ws.cell(row=9, column=1, value="编制/日期：\nAuthor/Date")
        ws.cell(row=9, column=3, value=project.get("author_date", datetime.now().strftime("%Y.%m")))
        
        ws.cell(row=10, column=1, value="审核/日期：\nReview/Date")
        ws.cell(row=10, column=3, value=project.get("review_date", ""))
        
        ws.cell(row=11, column=1, value="会签/日期：\nSignature/Date")
        ws.cell(row=11, column=3, value=project.get("signature_date", ""))
        
        ws.cell(row=12, column=1, value="批准/日期：\nApprove/Date")
        ws.cell(row=12, column=3, value=project.get("approve_date", ""))
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['F'].width = 50

    def _create_definitions_sheet(self, wb: Workbook, project: dict):
        """Create 1-相关定义 sheet with definitions."""
        ws = wb.create_sheet("1-相关定义", 1)

        project_name = project.get("name", "TARA分析项目")
        
        # Title
        title_cell = ws.cell(row=1, column=1, value=f"{project_name} TARA分析报告 - 相关定义")
        title_cell.font = Font(bold=True, size=14, color="2F5496")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        # Section 1: Functional Description
        row = 3
        ws.cell(row=row, column=1, value="1. 功能描述 Functional Description").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        row = 4
        func_desc = project.get("description", 
            "车载信息娱乐系统(In-Vehicle Infotainment, IVI)是一种集成多媒体娱乐、导航、车辆信息显示、通信和车辆控制功能于一体的车载电子系统。\n\n主要功能包括：\n• 多媒体播放：音乐、视频、图片等媒体文件的播放和管理\n• 导航定位：GPS定位、地图显示、路径规划和导航引导\n• 蓝牙通信：与手机连接实现免提通话和音乐传输\n• 车辆信息：显示车速、油耗、里程等车辆状态信息\n• 车辆控制：车窗、空调、座椅等舒适性功能的控制\n• 远程服务：OTA升级、远程诊断、紧急救援等服务")
        ws.cell(row=row, column=1, value=func_desc)
        ws.merge_cells(start_row=row, start_column=1, end_row=row+7, end_column=6)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Section 2: Item Boundary Diagram
        row = 12
        ws.cell(row=row, column=1, value="2. 项目边界 Item Boundary").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        row = 13
        # Get custom boundary diagram or use default
        boundary_diagram = project.get("boundary_diagram", """
┌─────────────────────────────────────────────────────────────────────────────────────┐
│                           车载信息娱乐系统 项目边界示意图                             │
│  ┌─────────────────────────────────────────────────────────────────────────────┐    │
│  │                         IVI 系统边界 (Scope Boundary)                        │    │
│  │  ┌──────────────────────────────────────────────────────────────────────┐   │    │
│  │  │                    硬件层 Hardware Layer                              │   │    │
│  │  │  ┌─────────┐ ┌─────────┐ ┌─────────┐ ┌─────────┐ ┌─────────┐        │   │    │
│  │  │  │   SOC   │ │   MCU   │ │ DDR RAM │ │ UFS ROM │ │  eMMC   │        │   │    │
│  │  │  └─────────┘ └─────────┘ └─────────┘ └─────────┘ └─────────┘        │   │    │
│  │  │  ┌─────────┐ ┌─────────┐ ┌─────────┐ ┌─────────┐                    │   │    │
│  │  │  │ BT模块  │ │ WIFI模块│ │以太网接口│ │ CAN接口 │                    │   │    │
│  │  │  └─────────┘ └─────────┘ └─────────┘ └─────────┘                    │   │    │
│  │  └──────────────────────────────────────────────────────────────────────┘   │    │
│  │                                 ↕ 内部通信                                   │    │
│  │  ┌──────────────────────────────────────────────────────────────────────┐   │    │
│  │  │                    软件层 Software Layer                              │   │    │
│  │  │  系统固件 │ 操作系统 │ 中间件 │ 应用软件 │ 安全服务 │ 配置数据        │   │    │
│  │  └──────────────────────────────────────────────────────────────────────┘   │    │
│  └─────────────────────────────────────────────────────────────────────────────┘    │
│                              ↕ 外部接口 External Interface                          │
│  ┌────────────────────────────┐  ┌────────────────────────────────────────────┐    │
│  │   外部接口 (项目边界)       │  │        外部系统 (项目范围外)                │    │
│  │  • TBox接口 (以太网/CAN)   │  │  • 云服务平台 (OTA服务器、TSP)             │    │
│  │  • OBD接口 (诊断)          │  │  • 用户手机 (蓝牙/WiFi连接)                │    │
│  │  • USB接口                 │  │  • 诊断设备                                │    │
│  │  • CAN总线接口             │  │  • 车身域控制器/网关                        │    │
│  └────────────────────────────┘  └────────────────────────────────────────────┘    │
└─────────────────────────────────────────────────────────────────────────────────────┘
""")
        ws.cell(row=row, column=1, value=boundary_diagram)
        ws.merge_cells(start_row=row, start_column=1, end_row=row+18, end_column=6)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=1).font = Font(name='Courier New', size=9)
        
        # Section 3: System Architecture Diagram
        row = 32
        ws.cell(row=row, column=1, value="3. 系统架构图 System Architecture").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        row = 33
        system_arch_diagram = project.get("system_architecture_diagram", """
┌─────────────────────────────────────────────────────────────────────────────────────┐
│                           IVI 系统架构图 System Architecture                         │
│                                                                                      │
│  ┌────────────────────────────────────────────────────────────────────────────────┐ │
│  │                           应用层 Application Layer                              │ │
│  │  ┌──────────┐ ┌──────────┐ ┌──────────┐ ┌──────────┐ ┌──────────┐ ┌──────────┐│ │
│  │  │ 导航APP  │ │ 多媒体APP │ │ 蓝牙APP  │ │ 车控APP  │ │ 设置APP  │ │ OTA APP  ││ │
│  │  └──────────┘ └──────────┘ └──────────┘ └──────────┘ └──────────┘ └──────────┘│ │
│  └────────────────────────────────────────────────────────────────────────────────┘ │
│                                        ↓                                            │
│  ┌────────────────────────────────────────────────────────────────────────────────┐ │
│  │                           中间件层 Middleware Layer                             │ │
│  │  ┌──────────┐ ┌──────────┐ ┌──────────┐ ┌──────────┐ ┌──────────┐ ┌──────────┐│ │
│  │  │ 音频管理 │ │ 显示管理 │ │ 网络管理 │ │ 安全服务 │ │ 诊断服务 │ │ OTA服务  ││ │
│  │  └──────────┘ └──────────┘ └──────────┘ └──────────┘ └──────────┘ └──────────┘│ │
│  └────────────────────────────────────────────────────────────────────────────────┘ │
│                                        ↓                                            │
│  ┌────────────────────────────────────────────────────────────────────────────────┐ │
│  │                           操作系统层 OS Layer                                   │ │
│  │                    Android Automotive OS / QNX / Linux                          │ │
│  └────────────────────────────────────────────────────────────────────────────────┘ │
│                                        ↓                                            │
│  ┌────────────────────────────────────────────────────────────────────────────────┐ │
│  │                           硬件抽象层 HAL Layer                                  │ │
│  │       Audio HAL │ Display HAL │ Camera HAL │ Vehicle HAL │ Sensors HAL         │ │
│  └────────────────────────────────────────────────────────────────────────────────┘ │
│                                        ↓                                            │
│  ┌────────────────────────────────────────────────────────────────────────────────┐ │
│  │                           硬件层 Hardware Layer                                 │ │
│  │  SOC │ MCU │ DDR │ UFS │ BT │ WIFI │ Ethernet │ CAN │ USB │ Display │ Audio    │ │
│  └────────────────────────────────────────────────────────────────────────────────┘ │
└─────────────────────────────────────────────────────────────────────────────────────┘
""")
        ws.cell(row=row, column=1, value=system_arch_diagram)
        ws.merge_cells(start_row=row, start_column=1, end_row=row+18, end_column=6)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=1).font = Font(name='Courier New', size=9)
        
        # Section 4: Software Architecture Diagram
        row = 52
        ws.cell(row=row, column=1, value="4. 软件架构图 Software Architecture").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        row = 53
        software_arch_diagram = project.get("software_architecture_diagram", """
┌─────────────────────────────────────────────────────────────────────────────────────┐
│                           IVI 软件架构图 Software Architecture                       │
│                                                                                      │
│  ┌────────────────────────────────────────────────────────────────────────────────┐ │
│  │  安全域 Secure World (ARM TrustZone)  │  普通域 Normal World                   │ │
│  │  ┌────────────────────────────────┐   │  ┌────────────────────────────────┐   │ │
│  │  │      TEE OS (OP-TEE)           │   │  │      Rich OS (Android/Linux)   │   │ │
│  │  │  ┌──────────────────────────┐  │   │  │  ┌──────────────────────────┐  │   │ │
│  │  │  │  Trusted Applications    │  │   │  │  │   User Applications      │  │   │ │
│  │  │  │  • Secure Boot           │  │   │  │  │  • 导航 Navigation       │  │   │ │
│  │  │  │  • Key Management        │  │   │  │  │  • 多媒体 Multimedia     │  │   │ │
│  │  │  │  • Crypto Services       │  │   │  │  │  • 蓝牙 Bluetooth        │  │   │ │
│  │  │  │  • Secure Storage        │  │   │  │  │  • 车控 Vehicle Control  │  │   │ │
│  │  │  │  • DRM Services          │  │   │  │  │  • OTA Update            │  │   │ │
│  │  │  └──────────────────────────┘  │   │  │  └──────────────────────────┘  │   │ │
│  │  │              ↑                 │   │  │              ↑                 │   │ │
│  │  │  ┌──────────────────────────┐  │   │  │  ┌──────────────────────────┐  │   │ │
│  │  │  │  TEE Internal API        │  │   │  │  │  System Services         │  │   │ │
│  │  │  │  (GlobalPlatform)        │  │   │  │  │  • Communication Mgr     │  │   │ │
│  │  │  └──────────────────────────┘  │   │  │  │  • Diagnostic Service    │  │   │ │
│  │  └────────────────────────────────┘   │  │  │  • Security Service      │  │   │ │
│  │              ↑                        │  │  └──────────────────────────┘  │   │ │
│  │  ┌────────────────────────────────┐   │  │              ↑                 │   │ │
│  │  │       HSM (Hardware)           │   │  │  ┌──────────────────────────┐  │   │ │
│  │  │  • Secure Key Storage          │   │  │  │  Drivers Layer           │  │   │ │
│  │  │  • Crypto Acceleration         │   │  │  │  CAN│ETH│USB│BT│WIFI│...│  │   │ │
│  │  └────────────────────────────────┘   │  │  └──────────────────────────┘  │   │ │
│  └───────────────────────────────────────┴──┴────────────────────────────────────┘ │
└─────────────────────────────────────────────────────────────────────────────────────┘
""")
        ws.cell(row=row, column=1, value=software_arch_diagram)
        ws.merge_cells(start_row=row, start_column=1, end_row=row+18, end_column=6)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=1).font = Font(name='Courier New', size=9)
        
        # Section 5: Item Assumptions
        row = 72
        ws.cell(row=row, column=1, value="5. 相关项假设 Item Assumptions").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        # Assumption headers
        row = 73
        ws.cell(row=row, column=1, value="假设编号\nAssumption ID").font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value="假设描述 Assumption Description").font = Font(bold=True, size=10)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        
        # Assumptions data
        assumptions = project.get("assumptions", [
            ("ASM-01", "IVI系统运行在独立的电子控制单元上，与车辆安全关键系统(如制动、转向)物理隔离"),
            ("ASM-02", "IVI系统与车身域控制器通过CAN网关进行通信，网关具备消息过滤和访问控制功能"),
            ("ASM-03", "TBox与IVI之间的通信链路为车内以太网(100BASE-T1)或CAN总线"),
            ("ASM-04", "普通用户不具备拆解IVI硬件的专业工具和能力，物理攻击需要专业设备"),
            ("ASM-05", "IVI系统具备安全启动(Secure Boot)功能，启动链完整性受硬件信任根保护"),
            ("ASM-06", "系统固件和应用软件由OEM统一管理和分发，具备代码签名机制"),
            ("ASM-07", "车辆使用期间，云服务器(TSP)持续在线提供OTA、远程诊断等服务"),
            ("ASM-08", "用户移动设备的安全性不在本分析范围内，假设手机可能被恶意软件感染"),
            ("ASM-09", "车辆网络(CAN/Ethernet)与外部网络之间存在网关隔离"),
            ("ASM-10", "诊断接口(OBD)在车辆正常使用时可被物理访问"),
        ])
        
        row = 74
        for asm_id, asm_desc in assumptions:
            ws.cell(row=row, column=1, value=asm_id)
            ws.cell(row=row, column=2, value=asm_desc)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 1
        
        # Section 6: Terminology
        row += 1
        ws.cell(row=row, column=1, value="6. 术语表 Terminology").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        # Terminology headers
        row += 1
        ws.cell(row=row, column=1, value="缩写\nAbbreviation").font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value="英文全称 English Full Name").font = Font(bold=True, size=10)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=6, value="中文全称 Chinese Name").font = Font(bold=True, size=10)
        
        # Terminology data
        terminology = project.get("terminology", [
            ("IVI", "In-Vehicle Infotainment", "车载信息娱乐系统"),
            ("TARA", "Threat Analysis and Risk Assessment", "威胁分析与风险评估"),
            ("SOC", "System on Chip", "片上系统"),
            ("MCU", "Microcontroller Unit", "微控制器单元"),
            ("DDR RAM", "Double Data Rate Random Access Memory", "双倍数据率随机存取存储器"),
            ("UFS ROM", "Universal Flash Storage", "通用闪存存储"),
            ("BT", "Bluetooth", "蓝牙"),
            ("WIFI", "Wireless Fidelity", "无线保真/无线网络"),
            ("TBox", "Telematics Box", "车载终端/远程信息处理器"),
            ("OBD", "On-Board Diagnostics", "车载诊断系统"),
            ("CAN", "Controller Area Network", "控制器局域网"),
            ("OTA", "Over-The-Air", "空中下载/远程升级"),
            ("TEE", "Trusted Execution Environment", "可信执行环境"),
            ("HSM", "Hardware Security Module", "硬件安全模块"),
            ("PKI", "Public Key Infrastructure", "公钥基础设施"),
            ("STRIDE", "Spoofing,Tampering,Repudiation,Information Disclosure,DoS,EoP", "欺骗、篡改、抵赖、信息泄露、拒绝服务、权限提升"),
            ("WP29", "World Forum for Harmonization of Vehicle Regulations", "世界车辆法规协调论坛"),
            ("SecOC", "Secure Onboard Communication", "安全车载通信"),
            ("UDS", "Unified Diagnostic Services", "统一诊断服务"),
            ("DoIP", "Diagnostics over IP", "基于IP的诊断"),
        ])
        
        row += 1
        for abbr, eng, chn in terminology:
            ws.cell(row=row, column=1, value=abbr)
            ws.cell(row=row, column=2, value=eng)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            ws.cell(row=row, column=6, value=chn)
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _create_asset_list_sheet(self, wb: Workbook, assets: list, project: dict):
        """Create 2-资产列表 sheet."""
        ws = wb.create_sheet("2-资产列表", 2)
        
        project_name = project.get("name", "TARA分析项目")

        # Title
        title_cell = ws.cell(row=1, column=1, value=f"{project_name}- 资产列表 Asset List")
        title_cell.font = Font(bold=True, size=14, color="2F5496")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

        # Header Group Row 3
        self._merge_and_style(ws, 3, 1, 3, 4, 
                             "Asset Identification 资产识别",
                             self.header_font, self.header_fill)
        self._merge_and_style(ws, 3, 5, 3, 10, 
                             "Cybersecurity Attributes 网络安全属性",
                             self.header_font, self.header_fill)

        # Header Row 4
        headers = [
            ("资产ID\nAsset ID", 10),
            ("资产名称\nAsset Name", 15),
            ("分类\nCategory", 12),
            ("备注\nRemarks", 40),
            ("真实性\nAuthenticity", 12),
            ("完整性\nIntegrity", 12),
            ("不可抵赖性\nNon-repudiation", 14),
            ("机密性\nConfidentiality", 12),
            ("可用性\nAvailability", 12),
            ("权限\nAuthorization", 12),
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            self._apply_cell_style(cell, self.subheader_font, self.subheader_fill, self.center_align)
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        for row_idx, asset in enumerate(assets, 5):
            asset_id = asset.get("id", f"A-{row_idx-4:03d}")
            if isinstance(asset_id, int):
                asset_id = f"A-{asset_id:03d}"
            
            # Get security attributes
            security_attrs = asset.get("security_attrs", asset.get("security_attributes", {})) or {}
            
            row_data = [
                asset_id,
                asset.get("name", ""),
                asset.get("category", asset.get("asset_type", "内部实体")),
                asset.get("description", asset.get("remarks", "")),
                "√" if security_attrs.get("authenticity") else "",
                "√" if security_attrs.get("integrity") else "",
                "√" if security_attrs.get("non_repudiation") else "",
                "√" if security_attrs.get("confidentiality") else "",
                "√" if security_attrs.get("availability") else "",
                "√" if security_attrs.get("authorization") else "",
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                self._apply_cell_style(cell, self.data_font, None, 
                                      self.center_align if col >= 5 else self.left_align)

    def _create_data_flow_sheet(self, wb: Workbook, assets: list, project: dict):
        """Create 3-数据流图 sheet."""
        ws = wb.create_sheet("3-数据流图", 3)
        
        project_name = project.get("name", "TARA分析项目")

        # Title
        title_cell = ws.cell(row=1, column=1, value=f"{project_name}- 数据流图 Data Flow Diagram")
        title_cell.font = Font(bold=True, size=14, color="2F5496")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

        # Data flow diagram
        diagram_text = project.get("data_flow_diagram", """
┌──────────────────────────────────────────────────────────────────────────────────────────────────────────────┐
│                                    IVI系统数据流图 Data Flow Diagram (Level 0)                                │
│                                                                                                               │
│     ┌─────────────┐                                                              ┌─────────────┐             │
│     │ 云服务器    │                                                              │  车身控制器  │             │
│     │ Cloud Server│                                                              │  BCM/Gateway│             │
│     │    (E1)     │                                                              │    (E6)     │             │
│     └──────┬──────┘                                                              └──────┬──────┘             │
│            │                                                                            │                     │
│            │ DF1: OTA升级包                                                             │ DF9: CAN车控指令     │
│            │ DF2: 远程配置                                                              │ DF10: 车辆状态      │
│            ↓                                                                            ↑                     │
│     ┌─────────────┐         DF3: 远程指令          ╔════════════════════════════════════════════════╗        │
│     │   TBox      │ ─────────────────────────────→ ║                                                ║        │
│     │ Telematics  │ ←───────────────────────────── ║           IVI 主系统                           ║        │
│     │    (E2)     │         DF4: 状态上报          ║         IVI Main System                        ║        │
│     └─────────────┘                                ║                                                ║        │
│                                                    ║  ┌────────────────────────────────────────┐    ║        │
│     ┌─────────────┐         DF5: 蓝牙音频/电话     ║  │         P1: SOC 主处理器                │    ║        │
│     │  用户手机   │ ─────────────────────────────→ ║  │         Main Processor                 │    ║        │
│     │ User Phone  │ ←───────────────────────────── ║  │  ┌────────┐  ┌────────┐  ┌────────┐   │    ║        │
│     │    (E3)     │         DF6: 联系人/通知       ║  │  │ D1:配置│  │ D2:密钥│  │D3:用户 │   │    ║ ──────→│
│     └─────────────┘                                ║  │  │  数据  │  │  存储  │  │  数据  │   │    ║        │
│                                                    ║  │  └────────┘  └────────┘  └────────┘   │    ║        │
│     ┌─────────────┐         DF7: 诊断请求          ║  │                    ↕                   │    ║        │
│     │  诊断设备   │ ─────────────────────────────→ ║  │         ┌────────────────────┐        │    ║        │
│     │ Diagnostic  │ ←───────────────────────────── ║  │         │   P2: MCU 安全控制  │        │    ║        │
│     │ Tool (E4)   │         DF8: 诊断响应          ║  │         │   Safety Controller │        │    ║        │
│     └─────────────┘                                ║  │         └────────────────────┘        │    ║        │
│                                                    ║  └────────────────────────────────────────┘    ║        │
│     ┌─────────────┐         DF11: 媒体文件         ║                                                ║        │
│     │  USB设备    │ ─────────────────────────────→ ║                                                ║        │
│     │ USB Device  │                                ╚════════════════════════════════════════════════╝        │
│     │    (E5)     │                                                                                           │
│     └─────────────┘                                                                                           │
│                                                                                                               │
│  图例 Legend:  ═══ 信任边界 Trust Boundary    ───→ 数据流 Data Flow    [ ] 处理 Process    (Dx) 数据存储      │
└──────────────────────────────────────────────────────────────────────────────────────────────────────────────┘
""")
        ws.cell(row=3, column=1, value=diagram_text)
        ws.merge_cells(start_row=3, start_column=1, end_row=40, end_column=9)
        ws.cell(row=3, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=3, column=1).font = Font(name='Courier New', size=9)

        # Adjust column widths
        ws.column_dimensions['A'].width = 120
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_attack_tree_sheet(self, wb: Workbook, threats: list, project: dict):
        """Create 4-攻击树图 sheet."""
        ws = wb.create_sheet("4-攻击树图", 4)
        
        project_name = project.get("name", "TARA分析项目")

        # Title
        title_cell = ws.cell(row=1, column=1, value=f"{project_name} - 攻击树分析 Attack Tree Analysis")
        title_cell.font = Font(bold=True, size=14, color="2F5496")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        # Default attack trees (4 diagrams matching the sample)
        default_attack_trees = [
            {
                "name": "攻击树1 Attack Tree 1: 远程入侵IVI系统 Remote Compromise IVI",
                "diagram": """                                        ┌─────────────────────────────────┐
                                        │     【攻击目标 Attack Goal】     │
                                        │     远程入侵IVI系统              │
                                        │     Remote Compromise IVI       │
                                        └────────────────┬────────────────┘
                                                         │
                     ┌───────────────────────────────────┼───────────────────────────────────┐
                     │                                   │                                   │
         ┌───────────┴───────────┐           ┌───────────┴───────────┐           ┌───────────┴───────────┐
         │ [OR] 通过TBox入侵     │           │ [OR] 通过WiFi入侵     │           │ [OR] 通过云服务入侵   │
         │ Via TBox (AT1.1)      │           │ Via WiFi (AT1.2)      │           │ Via Cloud (AT1.3)     │
         └───────────┬───────────┘           └───────────┬───────────┘           └───────────┬───────────┘
                     │                                   │                                   │
     ┌───────────────┼───────────────┐       ┌───────────┼───────────┐       ┌───────────────┼───────────────┐
     │               │               │       │           │           │       │               │               │
┌────┴────┐    ┌─────┴─────┐   ┌─────┴─────┐ │     ┌─────┴─────┐     │ ┌─────┴─────┐   ┌─────┴─────┐   ┌─────┴─────┐
│ 利用4G  │    │ 伪造远程  │   │ TBox固件  │ │     │ WiFi协议  │     │ │ 伪造OTA   │   │ 中间人    │   │ 云服务器  │
│ 协议漏洞│    │ 控制指令  │   │ 漏洞利用  │ │     │ 漏洞利用  │     │ │ 升级包    │   │ 攻击MITM  │   │ 被攻陷    │
│(1.1.1)  │    │ (1.1.2)   │   │ (1.1.3)   │ │     │ (1.2.1)   │     │ │ (1.3.1)   │   │ (1.3.2)   │   │ (1.3.3)   │
└─────────┘    └───────────┘   └───────────┘ │     └───────────┘     │ └───────────┘   └───────────┘   └───────────┘
                                             │           │           │
                                       ┌─────┴─────┐┌────┴────┐┌─────┴─────┐
                                       │ 恶意WiFi  ││连接恶意 ││ 无线监听  │
                                       │ AP欺骗   ││ 热点    ││ 数据窃取  │
                                       │ (1.2.2)  ││(1.2.3)  ││ (1.2.4)   │
                                       └──────────┘└─────────┘└───────────┘
"""
            },
            {
                "name": "攻击树2 Attack Tree 2: 物理接口攻击IVI系统 Physical Attack on IVI",
                "diagram": """                                        ┌─────────────────────────────────┐
                                        │     【攻击目标 Attack Goal】     │
                                        │     物理接口攻击IVI系统          │
                                        │     Physical Attack on IVI      │
                                        └────────────────┬────────────────┘
                                                         │
                     ┌───────────────────────────────────┼───────────────────────────────────┐
                     │                                   │                                   │
         ┌───────────┴───────────┐           ┌───────────┴───────────┐           ┌───────────┴───────────┐
         │ [OR] OBD接口攻击      │           │ [OR] USB接口攻击      │           │ [OR] CAN总线攻击      │
         │ Via OBD (AT2.1)       │           │ Via USB (AT2.2)       │           │ Via CAN (AT2.3)       │
         └───────────┬───────────┘           └───────────┬───────────┘           └───────────┬───────────┘
                     │                                   │                                   │
     ┌───────────────┼───────────────┐       ┌───────────┼───────────┐       ┌───────────────┼───────────────┐
     │               │               │       │           │           │       │               │               │
┌────┴────┐    ┌─────┴─────┐   ┌─────┴─────┐┌────┴────┐┌────┴────┐┌────┴────┐┌────┴────┐   ┌─────┴─────┐   ┌─────┴─────┐
│非授权   │    │ 注入恶意  │   │ 读取敏感  ││恶意U盘  ││固件刷写 ││USB漏洞  ││CAN报文  │   │ CAN DoS   │   │ CAN中间人 │
│诊断访问 │    │ 诊断指令  │   │ 诊断数据  ││感染病毒 ││攻击     ││利用     ││注入伪造 │   │ 攻击      │   │ 攻击      │
│(2.1.1)  │    │ (2.1.2)   │   │ (2.1.3)   ││(2.2.1)  ││(2.2.2)  ││(2.2.3)  ││(2.3.1)  │   │ (2.3.2)   │   │ (2.3.3)   │
└─────────┘    └───────────┘   └───────────┘└─────────┘└─────────┘└─────────┘└─────────┘   └───────────┘   └───────────┘
"""
            },
            {
                "name": "攻击树3 Attack Tree 3: 窃取IVI系统敏感数据 Steal Sensitive Data",
                "diagram": """                                        ┌─────────────────────────────────┐
                                        │     【攻击目标 Attack Goal】     │
                                        │     窃取IVI系统敏感数据          │
                                        │     Steal Sensitive Data        │
                                        └────────────────┬────────────────┘
                                                         │
                     ┌───────────────────────────────────┼───────────────────────────────────┐
                     │                                   │                                   │
         ┌───────────┴───────────┐           ┌───────────┴───────────┐           ┌───────────┴───────────┐
         │ [OR] 窃取用户数据     │           │ [OR] 窃取车辆信息     │           │ [OR] 窃取安全凭证     │
         │ User Data (AT3.1)     │           │ Vehicle Info (AT3.2)  │           │ Credentials (AT3.3)   │
         └───────────┬───────────┘           └───────────┬───────────┘           └───────────┬───────────┘
                     │                                   │                                   │
     ┌───────────────┼───────────────┐       ┌───────────┼───────────┐       ┌───────────────┼───────────────┐
     │               │               │       │           │           │       │               │               │
┌────┴────┐    ┌─────┴─────┐   ┌─────┴─────┐┌────┴────┐┌────┴────┐┌────┴────┐┌────┴────┐   ┌─────┴─────┐   ┌─────┴─────┐
│导航轨迹 │    │ 通讯录    │   │ 账户信息  ││VIN信息  ││里程篡改 ││故障码   ││提取证书 │   │ 密钥窃取  │   │ 令牌劫持  │
│位置信息 │    │ 个人信息  │   │ 凭证窃取  ││泄露     ││读取     ││泄露     ││(3.3.1)  │   │ (3.3.2)   │   │ (3.3.3)   │
│(3.1.1)  │    │ (3.1.2)   │   │ (3.1.3)   ││(3.2.1)  ││(3.2.2)  ││(3.2.3)  │└─────────┘   └───────────┘   └───────────┘
└─────────┘    └───────────┘   └───────────┘└─────────┘└─────────┘└─────────┘
"""
            },
            {
                "name": "攻击树4 Attack Tree 4: 恶意控制车辆功能 Malicious Vehicle Control",
                "diagram": """                                        ┌─────────────────────────────────┐
                                        │     【攻击目标 Attack Goal】     │
                                        │     恶意控制车辆功能             │
                                        │     Malicious Vehicle Control   │
                                        └────────────────┬────────────────┘
                                                         │
                     ┌───────────────────────────────────┼───────────────────────────────────┐
                     │                                   │                                   │
         ┌───────────┴───────────┐           ┌───────────┴───────────┐           ┌───────────┴───────────┐
         │ [OR] 远程车控攻击     │           │ [OR] 本地车控攻击     │           │ [OR] 供应链攻击       │
         │ Remote Control(AT4.1)│           │ Local Control(AT4.2)  │           │ Supply Chain(AT4.3)   │
         └───────────┬───────────┘           └───────────┬───────────┘           └───────────┬───────────┘
                     │                                   │                                   │
     ┌───────────────┼───────────────┐       ┌───────────┼───────────┐       ┌───────────────┼───────────────┐
     │               │               │       │           │           │       │               │               │
┌────┴────┐    ┌─────┴─────┐   ┌─────┴─────┐┌────┴────┐┌────┴────┐┌────┴────┐┌────┴────┐   ┌─────┴─────┐   ┌─────┴─────┐
│伪造车控 │    │ 重放攻击  │   │ 会话劫持  ││APP提权  ││IVI越权  ││物理篡改 ││恶意固件 │   │ 后门植入  │   │ 供应商   │
│指令     │    │ 历史指令  │   │ 接管会话  ││车控     ││车控指令 ││MCU      ││植入     │   │ (4.3.2)   │   │ 被攻陷   │
│(4.1.1)  │    │ (4.1.2)   │   │ (4.1.3)   ││(4.2.1)  ││(4.2.2)  ││(4.2.3)  ││(4.3.1)  │   └───────────┘   │ (4.3.3)  │
└─────────┘    └───────────┘   └───────────┘└─────────┘└─────────┘└─────────┘└─────────┘                   └───────────┘
"""
            },
        ]
        
        attack_trees = project.get("attack_trees", default_attack_trees)
        
        row = 3
        for tree in attack_trees:
            ws.cell(row=row, column=1, value=tree.get("name", "攻击树")).font = Font(bold=True, size=11)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            
            ws.cell(row=row, column=1, value=tree.get("diagram", ""))
            ws.merge_cells(start_row=row, start_column=1, end_row=row+18, end_column=6)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row, column=1).font = Font(name='Courier New', size=9)
            row += 21

        # Adjust column width
        ws.column_dimensions['A'].width = 120

    def _create_tara_results_sheet(self, wb: Workbook, assets: list, threats: list, 
                                   measures: list, project: dict):
        """Create 5-TARA分析结果 sheet - Main analysis results with formulas.
        
        This sheet has 40 columns with proper Excel formulas for auto-calculated fields.
        """
        ws = wb.create_sheet("5-TARA分析结果", 5)
        
        project_name = project.get("name", "TARA分析项目")

        # Row 1: Title
        title_cell = ws.cell(row=1, column=1, value=f"{project_name}_TARA分析结果 TARA Analysis Results")
        title_cell.font = self.title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=40)

        # Row 3: Main header groups (matching sample exactly)
        header_groups = [
            (1, 6, "Asset Identification资产识别"),      # A-F
            (7, 11, "Threat & Damage Scenario\n威胁&损害场景"),  # G-K
            (12, 21, "Threat Analysis\n威胁分析"),        # L-U
            (22, 35, "Impact Analysis\n影响分析"),        # V-AI
            # Columns AJ(36) and AK(37) have no row 3 header - they're header in row 4 only
            (38, 40, "Risk Mitigation\n风险缓解"),        # AL-AN
        ]
        
        for start_col, end_col, title in header_groups:
            self._merge_and_style(ws, 3, start_col, 3, end_col, title,
                                 self.header_font, self.header_fill, self.center_align)
        
        # Apply border to unmerged header cells in row 3
        for col in [36, 37]:
            cell = ws.cell(row=3, column=col)
            cell.border = self.border
            cell.fill = self.header_fill

        # Row 4-5: Sub-headers (matching sample structure)
        # Headers that span row 4-5 (merged vertically)
        row45_merged_headers = [
            (1, "Asset\nID\n资产ID"),
            (2, "Asset Name\n资产名称"),
            (6, "Category\n分类"),
            (7, "Security Attributes\n安全属性"),
            (8, "STRIDE Model\nSTRIDE模型"),
            (9, "Potential Threat and Damage Scenario\n潜在威胁和损害场景"),
            (10, "Attack Path\n攻击路径"),
            (36, "Risk Level\n风险等级"),
            (37, "Risk Treatment Decision\n风险处置决策"),
            (38, "Security Goal\n安全目标"),
            (39, "Security Requirement\n安全需求"),
        ]
        
        for col, title in row45_merged_headers:
            self._merge_and_style(ws, 4, col, 5, col, title,
                                 self.subheader_font, self.subheader_fill, self.center_align)

        # Row 4 only headers (with sub-headers in row 5)
        row4_only_headers = [
            (3, 5, "细分类"),                          # C4:E4
            (12, 13, "Attack Vector(V)\n攻击向量"),    # L4:M4
            (14, 15, "Attack Complexity(C)\n攻击复杂度"),  # N4:O4
            (16, 17, "Privileges Required(P)\n权限要求"),  # P4:Q4
            (18, 19, "User Interaction(U)\n用户交互"),  # R4:S4
            (20, 21, "Attack Feasibility\n攻击可行性计算"),  # T4:U4
            (22, 24, "Safety\n安全"),                  # V4:X4
            (25, 27, "Financial\n经济"),               # Y4:AA4
            (28, 30, "Operational\n操作"),             # AB4:AD4
            (31, 33, "Privacy & Legislation\n隐私和法律"),  # AE4:AG4
            (34, 35, "Impact Level Calculation\n影响等级计算"),  # AH4:AI4
        ]
        
        for start_col, end_col, title in row4_only_headers:
            self._merge_and_style(ws, 4, start_col, 4, end_col, title,
                                 self.subheader_font, self.subheader_fill, self.center_align)
        
        # Row 4 single cell header (column 11 - 来源)
        cell = ws.cell(row=4, column=11, value="来源\n")
        self._apply_cell_style(cell, self.subheader_font, self.subheader_fill, self.center_align)
        
        # Row 4 column 40 (WP29 Source)
        cell = ws.cell(row=4, column=40, value="Source来源\n")
        self._apply_cell_style(cell, self.subheader_font, self.subheader_fill, self.center_align)

        # Row 5: Detail headers
        row5_headers = {
            3: "子领域一",
            4: "子领域二",
            5: "子领域三",
            11: "WP29威胁映射",
            12: "Time\n内容",
            13: "Value\n指标值",
            14: "Content\n内容",
            15: "Value\n指标值",
            16: "Level\n等级",
            17: "Value\n指标值",
            18: "Level\n等级",
            19: "Value\n指标值",
            20: "Calculation\n计算值",
            21: "Level\n等级",
            22: "Content\n内容",
            23: "Notes\n注释",
            24: "Value\n指标值",
            25: "Content\n内容",
            26: "Notes\n注释",
            27: "Value\n指标值",
            28: "Content\n内容",
            29: "Notes\n注释",
            30: "Value\n指标值",
            31: "Content\n内容",
            32: "Notes\n注释",
            33: "Value\n指标值",
            34: "Impact Calc\n影响计算",
            35: "Impact Level\n影响等级",
            40: "WP29 Control Mapping",
        }
        
        for col, title in row5_headers.items():
            cell = ws.cell(row=5, column=col, value=title)
            self._apply_cell_style(cell, self.detail_font, self.detail_fill, self.center_align)

        # Apply borders to all cells in rows 3-5 that don't have values
        for row_num in range(3, 6):
            for col in range(1, 41):
                cell = ws.cell(row=row_num, column=col)
                if cell.border != self.border:
                    cell.border = self.border

        # Prepare measures lookup
        measures_by_threat = {}
        for m in measures:
            tid = m.get("threat_id") or m.get("threat_risk_id")
            if tid:
                if tid not in measures_by_threat:
                    measures_by_threat[tid] = []
                measures_by_threat[tid].append(m)

        # Data rows with formulas
        # Track asset groups for merging cells
        row = 6
        asset_groups = []  # List of (start_row, end_row) tuples
        current_group_start = row
        
        for idx, threat in enumerate(threats):
            self._write_tara_data_row(ws, row, threat, assets, measures_by_threat)
            
            # Check if this is a new asset group or continuation
            asset_id = threat.get("asset_id_str", "")
            
            # Check if next threat is a continuation (empty asset_id_str)
            is_last = (idx == len(threats) - 1)
            next_is_continuation = False
            if not is_last:
                next_asset_id = threats[idx + 1].get("asset_id_str", "")
                next_is_continuation = (next_asset_id == "" or next_asset_id.strip() == "")
            
            # If next row is NOT a continuation or this is the last row, end the group
            if not next_is_continuation:
                if row > current_group_start:
                    # We have a group of multiple rows
                    asset_groups.append((current_group_start, row))
                current_group_start = row + 1
            
            row += 1
        
        # Merge cells for asset groups
        # Columns to merge: A(1), B(2), C(3), D(4), F(6) - but NOT E(5) which is category_sub3
        merge_columns = [1, 2, 3, 4]  # A, B, C, D
        
        for start_row, end_row in asset_groups:
            for col in merge_columns:
                if start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=col, 
                                   end_row=end_row, end_column=col)
                    # Ensure the merged cell has the right value and style
                    cell = ws.cell(row=start_row, column=col)
                    cell.border = self.border
                    cell.alignment = self.center_align if col > 2 else self.left_align
                    # Apply border to all cells in merge range
                    for r in range(start_row, end_row + 1):
                        ws.cell(row=r, column=col).border = self.border

        # Set column widths to match sample
        col_widths = {
            1: 8.0, 2: 12.0, 3: 10.0, 4: 13.0, 5: 13.0, 6: 14.9, 7: 27.0, 8: 12.3,
            9: 53.9, 10: 82.3, 11: 10.0, 12: 13.0, 13: 7.6, 14: 10.0, 15: 7.7,
            16: 10.0, 17: 8.0, 18: 10.0, 19: 7.9, 20: 8.0, 21: 10.0, 22: 14.0,
            23: 18.0, 24: 6.0, 25: 14.0, 26: 28.0, 27: 6.0, 28: 14.0, 29: 12.0,
            30: 6.0, 31: 14.0, 32: 23.7, 33: 6.0, 34: 8.0, 35: 10.0, 36: 13.0,
            37: 12.0, 38: 18.0, 39: 25.0, 40: 12.0
        }
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze panes
        ws.freeze_panes = 'A6'

    def _write_tara_data_row(self, ws, row: int, threat: dict, assets: list, 
                             measures_by_threat: dict):
        """Write a single data row with proper formulas.
        
        Input columns (from threat data):
        - A(1): Asset ID
        - B(2): Asset Name
        - C(3): Category subdivision 1
        - D(4): Category subdivision 2
        - E(5): Category subdivision 3
        - F(6): Category
        - G(7): Security Attributes
        - H(8): STRIDE Model
        - I(9): Threat and Damage Scenario
        - J(10): Attack Path
        - K(11): WP29 Mapping
        - L(12): Attack Vector content
        - N(14): Attack Complexity content
        - P(16): Privileges Required level
        - R(18): User Interaction level
        - V(22): Safety content
        - Y(25): Financial content
        - AB(28): Operational content
        - AE(31): Privacy content
        - AM(39): Security Requirement
        
        Auto-calculated columns (formulas):
        - M(13), O(15), Q(17), S(19): Value columns
        - T(20), U(21): Attack Feasibility
        - W(23), X(24), Z(26), AA(27), AC(29), AD(30), AF(32), AG(33): Impact notes/values
        - AH(34), AI(35): Impact calculation/level
        - AJ(36): Risk Level
        - AK(37): Risk Treatment
        - AL(38): Security Goal
        - AN(40): WP29 Control Mapping
        """
        # Get associated asset
        asset_id = threat.get("asset_id", "")
        asset = next((a for a in assets if a.get("id") == asset_id), {})
        
        # Input columns
        # A(1): Asset ID
        ws.cell(row=row, column=1, value=threat.get("asset_id_str", asset_id or ""))
        
        # B(2): Asset Name
        ws.cell(row=row, column=2, value=threat.get("asset_name", asset.get("name", "")))
        
        # C(3): Category subdivision 1
        ws.cell(row=row, column=3, value=threat.get("category_sub1", asset.get("category_sub1", "")))
        
        # D(4): Category subdivision 2
        ws.cell(row=row, column=4, value=threat.get("category_sub2", asset.get("category_sub2", "")))
        
        # E(5): Category subdivision 3
        ws.cell(row=row, column=5, value=threat.get("category_sub3", asset.get("category_sub3", "")))
        
        # F(6): Category
        ws.cell(row=row, column=6, value=threat.get("category", asset.get("category", "")))
        
        # G(7): Security Attributes
        ws.cell(row=row, column=7, value=threat.get("security_attribute", ""))
        
        # H(8): STRIDE Model
        ws.cell(row=row, column=8, value=threat.get("stride_model", ""))
        
        # I(9): Threat and Damage Scenario
        ws.cell(row=row, column=9, value=threat.get("threat_scenario", threat.get("description", "")))
        
        # J(10): Attack Path
        ws.cell(row=row, column=10, value=threat.get("attack_path", ""))
        
        # K(11): WP29 Mapping
        ws.cell(row=row, column=11, value=threat.get("wp29_mapping", ""))
        
        # L(12): Attack Vector content (Input)
        ws.cell(row=row, column=12, value=threat.get("attack_vector", ""))
        
        # M(13): Attack Vector value (Formula)
        ws.cell(row=row, column=13, 
                value=f'=IF(L{row}="网络",0.85,IF(L{row}="邻居",0.62,IF(L{row}="本地",0.55,IF(L{row}="物理",0.2))))')
        
        # N(14): Attack Complexity content (Input)
        ws.cell(row=row, column=14, value=threat.get("attack_complexity", ""))
        
        # O(15): Attack Complexity value (Formula)
        ws.cell(row=row, column=15, value=f'=IF(N{row}="低",0.77,IF(N{row}="高",0.44))')
        
        # P(16): Privileges Required level (Input)
        ws.cell(row=row, column=16, value=threat.get("privileges_required", ""))
        
        # Q(17): Privileges Required value (Formula)
        ws.cell(row=row, column=17, value=f'=IF(P{row}="无",0.85,IF(P{row}="低",0.62,IF(P{row}="高",0.27)))')
        
        # R(18): User Interaction level (Input)
        ws.cell(row=row, column=18, value=threat.get("user_interaction", ""))
        
        # S(19): User Interaction value (Formula)
        ws.cell(row=row, column=19, value=f'=IF(R{row}="不需要",0.85,IF(R{row}="需要",0.62))')
        
        # T(20): Attack Feasibility Calculation (Formula)
        ws.cell(row=row, column=20, value=f'=8.22*M{row}*O{row}*Q{row}*S{row}')
        
        # U(21): Attack Feasibility Level (Formula)
        ws.cell(row=row, column=21, 
                value=f'=IF(T{row}<=1.05,"很低",IF(T{row}<=1.99,"低",IF(T{row}<=2.95,"中",IF(T{row}<=3.89,"高"))))')
        
        # V(22): Safety content (Input)
        ws.cell(row=row, column=22, value=threat.get("safety_impact", ""))
        
        # W(23): Safety Notes (Formula)
        ws.cell(row=row, column=23, 
                value=f'=IF(V{row}="可忽略不计的","没有受伤",IF(V{row}="中等的","轻度和中度伤害",IF(V{row}="重大的","重伤和生命危险（可能幸存）",IF(V{row}="严重的","威胁生命的伤害（尚不确定生存），致命伤害"))))')
        
        # X(24): Safety Value (Formula)
        ws.cell(row=row, column=24, 
                value=f'=IF(V{row}="可忽略不计的",0,IF(V{row}="中等的",10,IF(V{row}="重大的",100,IF(V{row}="严重的",1000))))')
        
        # Y(25): Financial content (Input)
        ws.cell(row=row, column=25, value=threat.get("financial_impact", ""))
        
        # Z(26): Financial Notes (Formula)
        ws.cell(row=row, column=26, 
                value=f'=IF(Y{row}="可忽略不计的","财务损失不会产生任何影响，可忽略给后果或与利益相关者无关",IF(Y{row}="中等的","经济损失会带来不便的后果，受影响的利益相关者将能够用有限的资源来克服这些后果",IF(Y{row}="重大的","经济损失导致重大后果，受影响的利益相关者将能够克服",IF(Y{row}="严重的","经济损失导致灾难性后果，受影响的利益相关者可能无法克服"))))')
        
        # AA(27): Financial Value (Formula)
        ws.cell(row=row, column=27, 
                value=f'=IF(Y{row}="可忽略不计的",0,IF(Y{row}="中等的","10",IF(Y{row}="重大的",100,IF(Y{row}="严重的",1000))))')
        
        # AB(28): Operational content (Input)
        ws.cell(row=row, column=28, value=threat.get("operational_impact", ""))
        
        # AC(29): Operational Notes (Formula)
        ws.cell(row=row, column=29, 
                value=f'=IF(AB{row}="可忽略不计的","操作损坏不会导致车辆功能或性能下降，也不会导致性能下降",IF(AB{row}="中等的","操作损坏会导致车辆功能或性能的部分下降",IF(AB{row}="重大的","操作损坏导致车辆功能丧失",IF(AB{row}="严重的","操作损坏导致车辆从不希望的运行到无法运行的所有车辆不工作"))))')
        
        # AD(30): Operational Value (Formula)
        ws.cell(row=row, column=30, 
                value=f'=IF(AB{row}="可忽略不计的",0,IF(AB{row}="中等的",1,IF(AB{row}="重大的",10,IF(AB{row}="严重的",100))))')
        
        # AE(31): Privacy content (Input)
        ws.cell(row=row, column=31, value=threat.get("privacy_impact", ""))
        
        # AF(32): Privacy Notes (Formula)
        ws.cell(row=row, column=32, 
                value=f'=IF(AE{row}="可忽略不计的","隐私危害不会产生任何影响,也不会给道路使用者带来不便",IF(AE{row}="中等的","隐私危害给道路使用者带来极大的不便",IF(AE{row}="重大的","隐私危害会严重影响道路使用者",IF(AE{row}="严重的","隐私危害对道路使用者造成重大甚至不可逆转的影响"))))')
        
        # AG(33): Privacy Value (Formula)
        ws.cell(row=row, column=33, 
                value=f'=IF(AE{row}="可忽略不计的",0,IF(AE{row}="中等的",1,IF(AE{row}="重大的",10,IF(AE{row}="严重的",100))))')
        
        # AH(34): Impact Calculation (Formula)
        ws.cell(row=row, column=34, value=f'=SUM(X{row}+AA{row}+AD{row}+AG{row})')
        
        # AI(35): Impact Level (Formula)
        ws.cell(row=row, column=35, 
                value=f'=IF(AH{row}>=1000,"严重的",IF(AH{row}>=100,"重大的",IF(AH{row}>=20,"中等的",IF(AH{row}>=10,"微不足道","无影响"))))')
        
        # AJ(36): Risk Level (Formula - complex matrix)
        risk_formula = f'''=IF(AND(AI{row}="无影响",U{row}="无"),"QM",IF(AND(AI{row}="无影响",U{row}="很低"),"QM",IF(AND(AI{row}="无影响",U{row}="低"),"QM",IF(AND(AI{row}="无影响",U{row}="中"),"QM",IF(AND(AI{row}="无影响",U{row}="高"),"Low",IF(AND(AI{row}="微不足道",U{row}="无"),"QM",IF(AND(AI{row}="微不足道",U{row}="很低"),"Low",IF(AND(AI{row}="微不足道",U{row}="低"),"Low",IF(AND(AI{row}="微不足道",U{row}="中"),"Low",IF(AND(AI{row}="微不足道",U{row}="高"),"Medium",IF(AND(AI{row}="中等的",U{row}="无"),"QM",IF(AND(AI{row}="中等的",U{row}="很低"),"Low",IF(AND(AI{row}="中等的",U{row}="低"),"Medium",IF(AND(AI{row}="中等的",U{row}="中"),"Medium",IF(AND(AI{row}="中等的",U{row}="高"),"High",IF(AND(AI{row}="重大的",U{row}="无"),"QM",IF(AND(AI{row}="重大的",U{row}="很低"),"Low",IF(AND(AI{row}="重大的",U{row}="低"),"Medium",IF(AND(AI{row}="重大的",U{row}="中"),"High",IF(AND(AI{row}="重大的",U{row}="高"),"High",IF(AND(AI{row}="严重的",U{row}="无"),"Low",IF(AND(AI{row}="严重的",U{row}="很低"),"Medium",IF(AND(AI{row}="严重的",U{row}="低"),"High",IF(AND(AI{row}="严重的",U{row}="中"),"High",IF(AND(AI{row}="严重的",U{row}="高"),"Critical")))))))))))))))))))))))))'''
        ws.cell(row=row, column=36, value=risk_formula)
        
        # AK(37): Risk Treatment Decision (Formula)
        ws.cell(row=row, column=37, 
                value=f'=IF(OR(AJ{row}="QM",AJ{row}="Low"),"保留风险",IF(OR(AJ{row}="Medium",AJ{row}="Critical",AJ{row}="High"),"降低风险","保留风险"))')
        
        # AL(38): Security Goal (Formula)
        ws.cell(row=row, column=38, 
                value=f'=IF(AK{row}="保留风险","/",IF(OR(AK{row}="降低风险",AK{row}="转移风险",AK{row}="消除风险"),"应保护"&E{row}&"的"&G{row},"NA"))')
        
        # AM(39): Security Requirement (Input)
        ws.cell(row=row, column=39, value=threat.get("security_requirement", ""))
        
        # AN(40): WP29 Control Mapping (Formula)
        ws.cell(row=row, column=40, 
                value=f'=IF(H{row}="T篡改","M10",IF(H{row}="D拒绝服务","M13",IF(H{row}="I信息泄露","M12"&CHAR(10)&"M24",IF(H{row}="S欺骗","M10"&CHAR(10)&"M11",IF(H{row}="E权限提升","M8"&CHAR(10)&"M9",IF(H{row}="R抵赖","M9"&CHAR(10)&"M10"))))))')
        
        # Apply styles to all cells in this row
        for col in range(1, 41):
            cell = ws.cell(row=row, column=col)
            cell.border = self.border
            cell.font = self.data_font
            cell.alignment = self.center_align if col > 2 else self.left_align
