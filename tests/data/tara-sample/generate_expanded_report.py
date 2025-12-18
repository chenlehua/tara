#!/usr/bin/env python3
"""
Generate Expanded TARA Analysis Report for MY25 EV Platform IVI System
=======================================================================

This script generates a comprehensive Excel report for the MY25 EV platform
central control host (中控主机) TARA analysis with 45+ assets, 30 threats,
and 43 control measures.
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class TARAReportGenerator:
    """Generate comprehensive TARA analysis Excel report."""

    def __init__(self):
        # Define styles
        self.title_font = Font(bold=True, size=16, color="FFFFFF")
        self.title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        self.header_font = Font(bold=True, size=10, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        self.header_fill_alt = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        self.subheader_font = Font(bold=True, size=9)
        self.subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        
        self.data_font = Font(size=9)
        self.alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Risk colors
        self.risk_fills = {
            'CAL-4': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            'CAL-3': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'CAL-2': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            'CAL-1': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        }

    def _apply_cell_style(self, cell, font=None, fill=None, alignment=None, border=True):
        """Apply styles to a cell."""
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = self.border

    def _merge_and_style(self, ws, start_row, start_col, end_row, end_col, value, font=None, fill=None, alignment=None):
        """Merge cells and apply style."""
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col, value=value)
        self._apply_cell_style(cell, font, fill, alignment or self.center_align)
        
        # Apply border to all merged cells
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = self.border

    def generate(self, assets_file, threats_file, measures_file, output_file):
        """Generate the complete Excel report."""
        # Load data
        with open(assets_file, 'r', encoding='utf-8') as f:
            assets_data = json.load(f)
        with open(threats_file, 'r', encoding='utf-8') as f:
            threats_data = json.load(f)
        with open(measures_file, 'r', encoding='utf-8') as f:
            measures_data = json.load(f)

        project = assets_data.get('project', {})
        assets = assets_data.get('assets', [])
        threats = threats_data.get('threats', [])
        measures = measures_data.get('control_measures', [])

        # Create workbook
        wb = Workbook()

        # Create sheets
        self._create_cover_sheet(wb, project)
        self._create_definitions_sheet(wb, project)
        self._create_asset_list_sheet(wb, assets)
        self._create_threat_analysis_sheet(wb, threats, assets, measures)
        self._create_tara_results_sheet(wb, assets, threats, measures)
        self._create_measures_sheet(wb, measures, threats)
        self._create_risk_summary_sheet(wb, threats)

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Save
        wb.save(output_file)
        print(f"✓ Report generated: {output_file}")
        print(f"  - Assets: {len(assets)}")
        print(f"  - Threats: {len(threats)}")
        print(f"  - Measures: {len(measures)}")

    def _create_cover_sheet(self, wb, project):
        """Create cover page sheet."""
        ws = wb.create_sheet("0-封面", 0)
        
        # Title
        self._merge_and_style(ws, 3, 2, 3, 10, 
                             "TARA 威胁分析与风险评估报告",
                             Font(bold=True, size=24, color="1F4E79"), None)
        
        self._merge_and_style(ws, 5, 2, 5, 10, 
                             "Threat Analysis and Risk Assessment Report",
                             Font(bold=True, size=16, color="5B9BD5"), None)
        
        # Project info
        row = 8
        info_items = [
            ("项目名称", project.get('name', 'MY25 EV平台中控主机 TARA分析')),
            ("车型平台", project.get('platform', 'E3-SOA平台')),
            ("车辆类型", project.get('vehicle_type', '纯电动乘用车')),
            ("分析范围", project.get('scope', '中控主机全系统')),
            ("参考标准", project.get('standard', 'ISO/SAE 21434:2021')),
            ("文档版本", "V2.0"),
            ("创建日期", datetime.now().strftime("%Y-%m-%d")),
            ("责任部门", project.get('owner', '网络安全工程部')),
        ]
        
        for label, value in info_items:
            ws.cell(row=row, column=2, value=label).font = Font(bold=True, size=12)
            ws.cell(row=row, column=4, value=value).font = Font(size=12)
            row += 1
        
        # Statistics
        row += 2
        ws.cell(row=row, column=2, value="分析统计").font = Font(bold=True, size=14)
        row += 1
        
        stats = [
            ("资产总数", "45"),
            ("威胁场景", "30"),
            ("控制措施", "43"),
            ("高风险项", "16 (CAL-4)"),
        ]
        for label, value in stats:
            ws.cell(row=row, column=2, value=label).font = Font(bold=True, size=11)
            ws.cell(row=row, column=4, value=value).font = Font(size=11, color="FF0000" if "CAL-4" in value else "000000")
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 50

    def _create_definitions_sheet(self, wb, project):
        """Create definitions sheet."""
        ws = wb.create_sheet("1-相关定义", 1)
        
        project_name = project.get("name", "MY25 EV平台中控主机")
        
        # Title
        self._merge_and_style(ws, 1, 1, 1, 12, 
                             f"{project_name} - 相关定义",
                             self.title_font, self.title_fill)
        
        # Sections
        row = 3
        
        # 1. Functional Description
        ws.cell(row=row, column=1, value="1. 功能描述 Functional Description").font = Font(bold=True, size=12)
        row += 1
        desc = project.get('description', '针对MY25年款纯电动汽车平台中控主机系统的全面威胁分析与风险评估')
        ws.cell(row=row, column=1, value=desc)
        ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=12)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        
        # 2. System Scope
        row = 9
        ws.cell(row=row, column=1, value="2. 分析范围 Analysis Scope").font = Font(bold=True, size=12)
        row += 1
        scope_items = [
            "• 主控单元：SoC主处理器、安全MCU、GPU/NPU等",
            "• 显示系统：中控屏、仪表盘、HUD、后排娱乐屏等",
            "• 音频系统：麦克风阵列、音频功放等",
            "• 通信模块：WiFi/蓝牙、5G/4G、GNSS、V2X、UWB、NFC等",
            "• 存储系统：eMMC、UFS、内存等",
            "• 网络设备：以太网交换机、CAN-FD网关等",
            "• 外部接口：USB、SD卡、OBD-II诊断接口等",
            "• 安全组件：HSM、SE安全元件等",
            "• 软件系统：Android Automotive、QNX Hypervisor等",
            "• 云端服务：语音助手、导航、OTA升级、远程车控等",
        ]
        for item in scope_items:
            ws.cell(row=row, column=1, value=item)
            row += 1
        
        # 3. Reference Standards
        row += 1
        ws.cell(row=row, column=1, value="3. 参考标准 Reference Standards").font = Font(bold=True, size=12)
        row += 1
        standards = [
            "• ISO/SAE 21434:2021 - Road vehicles — Cybersecurity engineering",
            "• UN R155 - Cyber Security Management System",
            "• UN R156 - Software Update Management System",
            "• GB/T 40857-2021 - 汽车网络安全测试方法",
            "• ISO 26262 - Functional Safety",
            "• AUTOSAR SecOC Specification",
        ]
        for std in standards:
            ws.cell(row=row, column=1, value=std)
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 100

    def _create_asset_list_sheet(self, wb, assets):
        """Create asset list sheet."""
        ws = wb.create_sheet("2-资产列表", 2)

        # Title
        self._merge_and_style(ws, 1, 1, 1, 12, 
                             "资产列表 Asset List",
                             self.title_font, self.title_fill)

        # Headers
        headers = [
            ("资产ID", 12),
            ("资产名称", 25),
            ("类型", 15),
            ("分类", 15),
            ("供应商", 12),
            ("描述", 50),
            ("接口", 25),
            ("机密性", 10),
            ("完整性", 10),
            ("可用性", 10),
            ("真实性", 10),
            ("授权", 10),
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.center_align)
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        for row_idx, asset in enumerate(assets, 4):
            sec_props = asset.get('security_properties', {})
            interfaces = asset.get('interfaces', [])
            interface_str = ', '.join([i.get('type', str(i)) if isinstance(i, dict) else str(i) for i in interfaces[:3]])
            
            row_data = [
                asset.get('id', ''),
                asset.get('name', ''),
                asset.get('type', ''),
                asset.get('category', ''),
                asset.get('supplier', '-'),
                asset.get('description', '')[:80],
                interface_str,
                sec_props.get('confidentiality', '-'),
                sec_props.get('integrity', '-'),
                sec_props.get('availability', '-'),
                '√' if sec_props.get('authenticity') else '',
                '√' if sec_props.get('authorization') else '',
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                fill = self.alt_fill if row_idx % 2 == 0 else None
                self._apply_cell_style(cell, self.data_font, fill, 
                                      self.center_align if col >= 8 else self.left_align)

    def _create_threat_analysis_sheet(self, wb, threats, assets, measures):
        """Create threat analysis sheet with STRIDE details."""
        ws = wb.create_sheet("3-威胁场景分析", 3)

        # Title
        self._merge_and_style(ws, 1, 1, 1, 15, 
                             "威胁场景分析 Threat Scenario Analysis",
                             self.title_font, self.title_fill)

        # Headers
        headers = [
            ("威胁ID", 12),
            ("资产", 20),
            ("STRIDE", 10),
            ("威胁名称", 30),
            ("威胁描述", 50),
            ("攻击向量", 12),
            ("攻击路径", 35),
            ("威胁代理", 15),
            ("攻击面", 15),
            ("可行性", 10),
            ("安全影响", 8),
            ("财务影响", 8),
            ("操作影响", 8),
            ("隐私影响", 8),
            ("风险等级", 10),
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.center_align)
            ws.column_dimensions[get_column_letter(col)].width = width

        # STRIDE mapping
        stride_map = {
            'S': 'S-欺骗', 'T': 'T-篡改', 'R': 'R-抵赖',
            'I': 'I-泄露', 'D': 'D-DoS', 'E': 'E-提权'
        }

        # Data rows
        for row_idx, threat in enumerate(threats, 4):
            impact = threat.get('impact', {})
            likelihood = threat.get('likelihood', {})
            
            row_data = [
                threat.get('id', ''),
                threat.get('asset_name', '')[:18],
                stride_map.get(threat.get('threat_type', 'T'), 'T-篡改'),
                threat.get('name', '')[:28],
                threat.get('description', '')[:80],
                threat.get('attack_vector', '-'),
                threat.get('attack_path', '-')[:33],
                threat.get('threat_agent', '-')[:13],
                threat.get('attack_surface', '-')[:13],
                likelihood.get('level', 'Medium'),
                f"S{impact.get('safety', 0)}",
                f"F{impact.get('financial', 0)}",
                f"O{impact.get('operational', 0)}",
                f"P{impact.get('privacy', 0)}",
                threat.get('risk_level', 'CAL-2'),
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                fill = self.alt_fill if row_idx % 2 == 0 else None
                
                # Color risk level
                if col == 15 and value in self.risk_fills:
                    fill = self.risk_fills[value]
                
                self._apply_cell_style(cell, self.data_font, fill, 
                                      self.center_align if col in [3, 6, 10, 11, 12, 13, 14, 15] else self.left_align)

        # Freeze panes
        ws.freeze_panes = 'A4'

    def _create_tara_results_sheet(self, wb, assets, threats, measures):
        """Create main TARA results sheet."""
        ws = wb.create_sheet("4-TARA分析结果", 4)

        # Title
        self._merge_and_style(ws, 1, 1, 1, 20, 
                             "TARA分析结果 TARA Analysis Results",
                             self.title_font, self.title_fill)

        # Headers
        headers = [
            ("资产ID", 10),
            ("资产名称", 18),
            ("分类", 12),
            ("安全属性", 12),
            ("STRIDE", 10),
            ("威胁名称", 25),
            ("攻击路径", 20),
            ("WP29 Ref", 8),
            ("攻击可行性", 10),
            ("安全S", 6),
            ("财务F", 6),
            ("操作O", 6),
            ("隐私P", 6),
            ("影响等级", 10),
            ("风险等级", 10),
            ("风险处置", 10),
            ("安全目标", 18),
            ("安全需求", 25),
            ("ISO条款", 8),
            ("状态", 8),
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.center_align)
            ws.column_dimensions[get_column_letter(col)].width = width

        # Build measures lookup
        measures_by_threat = {}
        for m in measures:
            for tid in m.get('threat_ids', []):
                if tid not in measures_by_threat:
                    measures_by_threat[tid] = []
                measures_by_threat[tid].append(m)

        # STRIDE and security attribute mapping
        stride_map = {
            'S': ('S-欺骗', '真实性'),
            'T': ('T-篡改', '完整性'),
            'R': ('R-抵赖', '不可抵赖'),
            'I': ('I-泄露', '机密性'),
            'D': ('D-DoS', '可用性'),
            'E': ('E-提权', '授权')
        }

        # Treatment mapping
        treatment_map = {
            'CAL-4': '降低',
            'CAL-3': '降低',
            'CAL-2': '降低/接受',
            'CAL-1': '接受'
        }

        # Data rows
        row = 4
        for threat in threats:
            asset_id = threat.get('asset_id', '')
            asset = next((a for a in assets if a.get('id') == asset_id), {})
            
            threat_type = threat.get('threat_type', 'T')
            stride_info = stride_map.get(threat_type, ('T-篡改', '完整性'))
            
            impact = threat.get('impact', {})
            risk_level = threat.get('risk_level', 'CAL-2')
            
            # Get measure for this threat
            threat_id = threat.get('id', '')
            threat_measures = measures_by_threat.get(threat_id, [])
            sec_goal = threat_measures[0].get('name', '-')[:16] if threat_measures else '-'
            sec_req = threat_measures[0].get('implementation', '-')[:23] if threat_measures else '-'
            
            # Impact level
            max_impact = max(impact.get('safety', 0), impact.get('financial', 0),
                           impact.get('operational', 0), impact.get('privacy', 0))
            impact_level = 'Severe' if max_impact >= 4 else 'Major' if max_impact >= 3 else 'Moderate' if max_impact >= 2 else 'Minor'
            
            row_data = [
                asset_id,
                asset.get('name', threat.get('asset_name', ''))[:16],
                asset.get('category', '-')[:10],
                stride_info[1],
                stride_info[0],
                threat.get('name', '')[:23],
                threat.get('attack_path', '-')[:18],
                threat.get('wp29_ref', '-'),
                threat.get('likelihood', {}).get('level', 'Medium')[:6],
                impact.get('safety', 0),
                impact.get('financial', 0),
                impact.get('operational', 0),
                impact.get('privacy', 0),
                impact_level[:8],
                risk_level,
                treatment_map.get(risk_level, '降低'),
                sec_goal,
                sec_req,
                threat.get('iso_clause', '-'),
                '已识别'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                fill = self.alt_fill if row % 2 == 0 else None
                
                # Color risk level
                if col == 15 and str(value) in self.risk_fills:
                    fill = self.risk_fills[str(value)]
                
                self._apply_cell_style(cell, self.data_font, fill, self.center_align)
            
            row += 1

        # Freeze panes
        ws.freeze_panes = 'A4'

    def _create_measures_sheet(self, wb, measures, threats):
        """Create control measures sheet."""
        ws = wb.create_sheet("5-控制措施", 5)

        # Title
        self._merge_and_style(ws, 1, 1, 1, 10, 
                             "控制措施 Control Measures",
                             self.title_font, self.title_fill)

        # Headers
        headers = [
            ("措施ID", 12),
            ("措施名称", 25),
            ("类型", 12),
            ("分类", 12),
            ("描述", 45),
            ("实施方法", 50),
            ("有效性", 10),
            ("优先级", 8),
            ("状态", 12),
            ("ISO参考", 12),
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.center_align)
            ws.column_dimensions[get_column_letter(col)].width = width

        # Status mapping
        status_map = {
            'implemented': '已实施',
            'in_progress': '进行中',
            'planned': '计划中'
        }

        # Data rows
        for row_idx, measure in enumerate(measures, 4):
            row_data = [
                measure.get('id', ''),
                measure.get('name', ''),
                measure.get('control_type', '-'),
                measure.get('category', '-'),
                measure.get('description', '')[:60],
                measure.get('implementation', '')[:70],
                measure.get('effectiveness', '-'),
                measure.get('priority', '-'),
                status_map.get(measure.get('status', ''), measure.get('status', '-')),
                measure.get('iso21434_ref', '-'),
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                fill = self.alt_fill if row_idx % 2 == 0 else None
                self._apply_cell_style(cell, self.data_font, fill, 
                                      self.center_align if col in [3, 7, 8, 9, 10] else self.left_align)

    def _create_risk_summary_sheet(self, wb, threats):
        """Create risk summary sheet."""
        ws = wb.create_sheet("6-风险统计", 6)

        # Title
        self._merge_and_style(ws, 1, 1, 1, 8, 
                             "风险统计分析 Risk Analysis Summary",
                             self.title_font, self.title_fill)

        # Risk distribution
        row = 3
        ws.cell(row=row, column=1, value="1. 风险等级分布").font = Font(bold=True, size=12)
        row += 1
        
        risk_counts = {'CAL-4': 0, 'CAL-3': 0, 'CAL-2': 0, 'CAL-1': 0}
        for threat in threats:
            risk_level = threat.get('risk_level', 'CAL-2')
            if risk_level in risk_counts:
                risk_counts[risk_level] += 1
        
        headers = ['风险等级', '描述', '数量', '占比']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.center_align)
        row += 1
        
        total = len(threats)
        risk_desc = {
            'CAL-4': '极高风险 - 必须降低',
            'CAL-3': '高风险 - 应当降低',
            'CAL-2': '中等风险 - 可降低/接受',
            'CAL-1': '低风险 - 可接受'
        }
        for risk_level in ['CAL-4', 'CAL-3', 'CAL-2', 'CAL-1']:
            count = risk_counts[risk_level]
            ratio = f"{count/total*100:.1f}%" if total > 0 else "0%"
            row_data = [risk_level, risk_desc[risk_level], count, ratio]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                fill = self.risk_fills.get(risk_level) if col == 1 else None
                self._apply_cell_style(cell, self.data_font, fill, self.center_align)
            row += 1

        # STRIDE distribution
        row += 2
        ws.cell(row=row, column=1, value="2. STRIDE威胁分布").font = Font(bold=True, size=12)
        row += 1
        
        stride_counts = {'S': 0, 'T': 0, 'R': 0, 'I': 0, 'D': 0, 'E': 0}
        for threat in threats:
            threat_type = threat.get('threat_type', 'T')
            if threat_type in stride_counts:
                stride_counts[threat_type] += 1
        
        headers = ['类型', '名称', '数量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.center_align)
        row += 1
        
        stride_names = {
            'S': 'Spoofing (欺骗)',
            'T': 'Tampering (篡改)',
            'R': 'Repudiation (抵赖)',
            'I': 'Information Disclosure (信息泄露)',
            'D': 'Denial of Service (拒绝服务)',
            'E': 'Elevation of Privilege (权限提升)'
        }
        for stride_type in ['S', 'T', 'R', 'I', 'D', 'E']:
            row_data = [stride_type, stride_names[stride_type], stride_counts[stride_type]]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                self._apply_cell_style(cell, self.data_font, None, self.center_align if col != 2 else self.left_align)
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10


def main():
    """Generate the expanded TARA report."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    assets_file = os.path.join(script_dir, 'MY25_IVI_expanded_assets.json')
    threats_file = os.path.join(script_dir, 'MY25_IVI_expanded_threats.json')
    measures_file = os.path.join(script_dir, 'MY25_IVI_expanded_measures.json')
    output_file = os.path.join(script_dir, 'MY25_EV平台中控主机_TARA分析报告_扩展版.xlsx')
    
    print("=" * 60)
    print("MY25 EV平台中控主机 TARA分析报告生成")
    print("=" * 60)
    
    generator = TARAReportGenerator()
    generator.generate(assets_file, threats_file, measures_file, output_file)
    
    print("\n" + "=" * 60)
    print("报告生成完成!")
    print("=" * 60)


if __name__ == "__main__":
    main()
