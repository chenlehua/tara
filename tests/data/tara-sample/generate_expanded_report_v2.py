#!/usr/bin/env python3
"""
Generate Expanded TARA Analysis Report for MY25 EV Platform IVI System
=======================================================================

This script generates an expanded Excel report maintaining the exact same
format as the original MY25 EV平台中控主机_TARA分析报告.xlsx file.
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class TARAReportGenerator:
    """Generate TARA analysis Excel report with same format as original."""

    def __init__(self):
        # Define styles matching original
        self.title_font = Font(bold=True, size=16, color="FFFFFF")
        self.title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        self.header_font = Font(bold=True, size=10, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        self.header_fill_alt = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        self.subheader_font = Font(bold=True, size=9)
        self.subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        
        self.data_font = Font(size=9)
        self.alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Risk colors
        self.risk_fills = {
            'CAL-4': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            'CAL-3': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'CAL-2': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            'CAL-1': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
        }

        # Expanded assets for MY25 EV IVI System
        self.assets = self._get_expanded_assets()
        self.threats = self._get_expanded_threats()

    def _get_expanded_assets(self):
        """Get expanded asset list for IVI system."""
        return [
            # Internal entities - 内部实体
            {"id": "P001", "name": "SOC", "category": "内部实体", 
             "desc": "主处理器芯片，运行Android Automotive OS，高通SA8295P",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": ""},
            {"id": "P002", "name": "MCU", "category": "内部实体",
             "desc": "安全微控制器，负责车控通信与功能安全，NXP S32K344",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "", "avail": "√", "authz": "√"},
            {"id": "P003", "name": "GPU", "category": "内部实体",
             "desc": "图形处理单元，负责多屏渲染和AR-HUD显示",
             "auth": "", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": ""},
            {"id": "P004", "name": "NPU", "category": "内部实体",
             "desc": "神经网络处理器，用于DMS/OMS驾驶员监控和语音AI",
             "auth": "", "integ": "√", "nonrep": "", "conf": "√", "avail": "√", "authz": ""},
            {"id": "P005", "name": "HSM安全模块", "category": "内部实体",
             "desc": "硬件安全模块，负责密钥存储和密码学运算",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "√", "authz": "√"},
            {"id": "P006", "name": "SE安全元件", "category": "内部实体",
             "desc": "安全元件，存储数字钥匙和支付凭证",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "", "authz": "√"},
            
            # Firmware and applications - 固件及应用
            {"id": "P011", "name": "媒体源", "category": "固件及应用",
             "desc": "负责接收视频及人脸信息，发送至存储和显示",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": ""},
            {"id": "P012", "name": "Android OS", "category": "固件及应用",
             "desc": "Android Automotive操作系统",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": "√"},
            {"id": "P013", "name": "QNX Hypervisor", "category": "固件及应用",
             "desc": "虚拟化管理程序，多OS隔离运行",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "√", "authz": "√"},
            {"id": "P014", "name": "Bootloader", "category": "固件及应用",
             "desc": "安全启动引导程序",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "", "avail": "√", "authz": ""},
            {"id": "P015", "name": "语音助手服务", "category": "固件及应用",
             "desc": "智能语音控制服务",
             "auth": "√", "integ": "", "nonrep": "", "conf": "√", "avail": "", "authz": "√"},
            {"id": "P016", "name": "导航应用", "category": "固件及应用",
             "desc": "高精度导航服务应用",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": ""},
            {"id": "P017", "name": "OTA升级客户端", "category": "固件及应用",
             "desc": "远程软件升级服务客户端",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "", "avail": "√", "authz": "√"},
            {"id": "P018", "name": "远程车控服务", "category": "固件及应用",
             "desc": "手机APP远程控制服务",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "", "authz": "√"},
            {"id": "P019", "name": "应用商店", "category": "固件及应用",
             "desc": "车载应用商店服务",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "", "avail": "", "authz": "√"},
            
            # External entities - 外部实体
            {"id": "E001", "name": "TBox", "category": "外部实体",
             "desc": "5G远程通信单元，提供车联网连接",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "√", "authz": ""},
            {"id": "E002", "name": "WiFi模块", "category": "外部实体",
             "desc": "WiFi 6E无线通信模块",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "√", "avail": "", "authz": ""},
            {"id": "E003", "name": "蓝牙模块", "category": "外部实体",
             "desc": "蓝牙5.3通信模块，支持LE Audio",
             "auth": "√", "integ": "", "nonrep": "", "conf": "√", "avail": "", "authz": ""},
            {"id": "E004", "name": "以太网接口", "category": "外部实体",
             "desc": "车载以太网通信接口，支持TSN",
             "auth": "", "integ": "", "nonrep": "", "conf": "", "avail": "", "authz": "√"},
            {"id": "E005", "name": "UART接口", "category": "外部实体",
             "desc": "调试通讯接口",
             "auth": "", "integ": "", "nonrep": "", "conf": "", "avail": "", "authz": "√"},
            {"id": "E006", "name": "USB接口", "category": "外部实体",
             "desc": "USB Type-C接口，支持CarPlay/Android Auto",
             "auth": "", "integ": "√", "nonrep": "", "conf": "√", "avail": "", "authz": ""},
            {"id": "E007", "name": "CAN网关", "category": "外部实体",
             "desc": "CAN-FD总线网关，域间通信",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": "√"},
            {"id": "E008", "name": "OBD诊断接口", "category": "外部实体",
             "desc": "车载诊断接口，支持UDS协议",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "", "avail": "", "authz": "√"},
            {"id": "E009", "name": "外置功放", "category": "外部实体",
             "desc": "高保真音频功放模块",
             "auth": "√", "integ": "", "nonrep": "", "conf": "", "avail": "", "authz": ""},
            {"id": "E010", "name": "GNSS模块", "category": "外部实体",
             "desc": "多频GNSS定位模块，GPS/BDS/GLONASS",
             "auth": "", "integ": "√", "nonrep": "", "conf": "√", "avail": "√", "authz": ""},
            {"id": "E011", "name": "V2X模块", "category": "外部实体",
             "desc": "C-V2X车路协同通信模块",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "", "avail": "√", "authz": ""},
            {"id": "E012", "name": "UWB模块", "category": "外部实体",
             "desc": "超宽带数字钥匙模块",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "", "authz": "√"},
            {"id": "E013", "name": "NFC模块", "category": "外部实体",
             "desc": "近场通信模块，手机/卡片钥匙",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "", "avail": "", "authz": "√"},
            {"id": "E014", "name": "DMS摄像头", "category": "外部实体",
             "desc": "驾驶员监控摄像头，红外夜视",
             "auth": "", "integ": "√", "nonrep": "", "conf": "√", "avail": "√", "authz": ""},
            {"id": "E015", "name": "OMS摄像头", "category": "外部实体",
             "desc": "乘客监控摄像头，儿童检测",
             "auth": "", "integ": "", "nonrep": "", "conf": "√", "avail": "", "authz": ""},
            {"id": "E016", "name": "麦克风阵列", "category": "外部实体",
             "desc": "6麦克风阵列，声源定位和降噪",
             "auth": "", "integ": "", "nonrep": "", "conf": "√", "avail": "", "authz": ""},
            {"id": "E017", "name": "中控显示屏", "category": "外部实体",
             "desc": "15.6寸OLED触控显示屏",
             "auth": "", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": ""},
            {"id": "E018", "name": "仪表显示屏", "category": "外部实体",
             "desc": "12.3寸全液晶仪表盘",
             "auth": "", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": ""},
            {"id": "E019", "name": "HUD抬头显示", "category": "外部实体",
             "desc": "AR-HUD增强现实抬头显示",
             "auth": "", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": ""},
            {"id": "E020", "name": "后排娱乐屏", "category": "外部实体",
             "desc": "双屏后排娱乐系统",
             "auth": "", "integ": "", "nonrep": "", "conf": "", "avail": "", "authz": ""},
            
            # Data storage - 数据存储
            {"id": "S001", "name": "eMMC存储", "category": "数据存储",
             "desc": "256GB嵌入式存储，系统和应用",
             "auth": "", "integ": "√", "nonrep": "", "conf": "√", "avail": "√", "authz": "√"},
            {"id": "S002", "name": "UFS存储", "category": "数据存储",
             "desc": "512GB高速存储，用户数据和地图",
             "auth": "", "integ": "√", "nonrep": "", "conf": "√", "avail": "", "authz": "√"},
            {"id": "S003", "name": "DDR内存", "category": "数据存储",
             "desc": "16GB LPDDR5内存",
             "auth": "", "integ": "√", "nonrep": "", "conf": "√", "avail": "√", "authz": ""},
            {"id": "S004", "name": "密钥存储", "category": "数据存储",
             "desc": "HSM内部密钥存储区",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "√", "authz": "√"},
            {"id": "S005", "name": "OTA升级包", "category": "数据存储",
             "desc": "存储在DDR中的固件升级包",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "√", "avail": "", "authz": "√"},
            {"id": "S006", "name": "用户配置", "category": "数据存储",
             "desc": "用户偏好设置和账户信息",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "√", "avail": "", "authz": "√"},
            {"id": "S007", "name": "导航历史", "category": "数据存储",
             "desc": "导航轨迹和收藏地点",
             "auth": "", "integ": "", "nonrep": "", "conf": "√", "avail": "", "authz": "√"},
            {"id": "S008", "name": "行车记录", "category": "数据存储",
             "desc": "行车视频和事件数据",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "", "authz": ""},
            
            # Data flow - 数据流
            {"id": "D001", "name": "远程控制指令", "category": "数据流",
             "desc": "云端→TBox→IVI",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "", "authz": "√"},
            {"id": "D002", "name": "OTA升级数据", "category": "数据流",
             "desc": "云端→TBox→IVI固件",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "", "avail": "√", "authz": "√"},
            {"id": "D003", "name": "CAN车控报文", "category": "数据流",
             "desc": "IVI↔MCU↔CAN网关",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "", "avail": "√", "authz": ""},
            {"id": "D004", "name": "蓝牙音频流", "category": "数据流",
             "desc": "手机→蓝牙→IVI→功放",
             "auth": "√", "integ": "", "nonrep": "", "conf": "√", "avail": "", "authz": ""},
            {"id": "D005", "name": "诊断数据", "category": "数据流",
             "desc": "诊断工具↔OBD↔IVI",
             "auth": "√", "integ": "√", "nonrep": "", "conf": "√", "avail": "", "authz": "√"},
            {"id": "D006", "name": "V2X消息", "category": "数据流",
             "desc": "路侧单元/其他车辆↔V2X模块",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "", "avail": "√", "authz": ""},
            {"id": "D007", "name": "数字钥匙信号", "category": "数据流",
             "desc": "手机→UWB/NFC→SE→车身控制",
             "auth": "√", "integ": "√", "nonrep": "√", "conf": "√", "avail": "", "authz": "√"},
            {"id": "D008", "name": "DMS视频流", "category": "数据流",
             "desc": "DMS摄像头→NPU→驾驶员状态",
             "auth": "", "integ": "√", "nonrep": "", "conf": "√", "avail": "√", "authz": ""},
        ]

    def _get_expanded_threats(self):
        """Get expanded threat list for TARA analysis."""
        return [
            # SOC threats
            {"asset_id": "P001", "asset_name": "SOC", "subdomain1": "车载多媒体", "subdomain2": "系统实体", "subdomain3": "N/A",
             "category": "内部实体", "sec_attr": "Authenticity\n真实性", "stride": "S欺骗",
             "threat": "攻击者通过物理手段或远程漏洞伪造SOC模块身份，获取系统控制权",
             "attack_path": "1.物理拆解获取芯片调试接口\n2.利用漏洞注入恶意代码\n3.伪造合法进程执行",
             "wp29_ref": "4.1\n5.1\n6.1", "vector": "本地", "complexity": "低", "privilege": "低", "interaction": "不需要",
             "safety": "S3严重", "safety_note": "可能导致驾驶安全功能失效", "safety_val": "3",
             "financial": "F2中", "financial_note": "维修和召回成本", "financial_val": "2",
             "operational": "O3严重", "operational_note": "系统完全失控", "operational_val": "3",
             "privacy": "P2中", "privacy_note": "用户数据泄露风险", "privacy_val": "2",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-01: SOC身份认证",
             "requirement": "实施安全启动和运行时完整性验证", "source": "M-001"},
            
            {"asset_id": "P001", "asset_name": "SOC", "subdomain1": "车载多媒体", "subdomain2": "系统实体", "subdomain3": "N/A",
             "category": "内部实体", "sec_attr": "Availability\n可用性", "stride": "D拒绝服务",
             "threat": "攻击者通过资源耗尽或恶意负载导致SOC服务不可用",
             "attack_path": "1.发送大量恶意请求导致CPU过载\n2.内存耗尽攻击\n3.恶意应用占用资源",
             "wp29_ref": "24.1\n8.1", "vector": "本地", "complexity": "低", "privilege": "低", "interaction": "不需要",
             "safety": "S2中", "safety_note": "关键显示可能中断", "safety_val": "2",
             "financial": "F1低", "financial_note": "临时服务中断", "financial_val": "1",
             "operational": "O3严重", "operational_note": "IVI功能不可用", "operational_val": "3",
             "privacy": "P0无", "privacy_note": "不涉及隐私", "privacy_val": "0",
             "risk_level": "CAL-3", "treatment": "降低", "goal": "SC-02: 资源访问控制",
             "requirement": "实施应用沙箱和资源配额限制", "source": "M-002"},
            
            {"asset_id": "P001", "asset_name": "SOC", "subdomain1": "车载多媒体", "subdomain2": "系统实体", "subdomain3": "N/A",
             "category": "内部实体", "sec_attr": "Authorization\n权限", "stride": "E提权",
             "threat": "攻击者利用Android内核漏洞获取Root权限",
             "attack_path": "1.恶意应用安装\n2.内核漏洞利用\n3.SELinux绕过获取Root",
             "wp29_ref": "4.3.5", "vector": "网络", "complexity": "高", "privilege": "低", "interaction": "需要",
             "safety": "S3严重", "safety_note": "可控制所有车辆功能", "safety_val": "3",
             "financial": "F3严重", "financial_note": "品牌声誉损失", "financial_val": "3",
             "operational": "O3严重", "operational_note": "系统完全被控", "operational_val": "3",
             "privacy": "P3严重", "privacy_note": "所有数据可被访问", "privacy_val": "3",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-03: 权限提升防护",
             "requirement": "启用SELinux强制模式，实施应用签名验证", "source": "M-003"},
            
            # MCU threats  
            {"asset_id": "P002", "asset_name": "MCU", "subdomain1": "车载多媒体", "subdomain2": "安全控制", "subdomain3": "N/A",
             "category": "内部实体", "sec_attr": "Integrity\n完整性", "stride": "T篡改",
             "threat": "攻击者通过故障注入绕过安全启动，加载恶意固件",
             "attack_path": "1.电压毛刺注入\n2.时钟故障攻击\n3.绕过签名验证加载恶意代码",
             "wp29_ref": "5.1.1", "vector": "物理", "complexity": "高", "privilege": "无", "interaction": "不需要",
             "safety": "S3严重", "safety_note": "车控功能被篡改", "safety_val": "3",
             "financial": "F3严重", "financial_note": "安全召回成本", "financial_val": "3",
             "operational": "O3严重", "operational_note": "安全功能失效", "operational_val": "3",
             "privacy": "P1低", "privacy_note": "密钥可能泄露", "privacy_val": "1",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-04: 安全启动加固",
             "requirement": "实施防故障注入措施，多重验证机制", "source": "M-004"},
            
            # HSM threats
            {"asset_id": "P005", "asset_name": "HSM安全模块", "subdomain1": "车载多媒体", "subdomain2": "安全组件", "subdomain3": "N/A",
             "category": "内部实体", "sec_attr": "Confidentiality\n机密性", "stride": "I信息泄露",
             "threat": "通过侧信道攻击提取HSM中存储的密钥",
             "attack_path": "1.功耗分析(DPA)攻击\n2.电磁分析(EMA)攻击\n3.时序攻击",
             "wp29_ref": "4.3.3", "vector": "物理", "complexity": "高", "privilege": "无", "interaction": "不需要",
             "safety": "S1低", "safety_note": "间接影响", "safety_val": "1",
             "financial": "F3严重", "financial_note": "系统安全被破坏", "financial_val": "3",
             "operational": "O2中", "operational_note": "需要密钥更换", "operational_val": "2",
             "privacy": "P3严重", "privacy_note": "通信可被解密", "privacy_val": "3",
             "risk_level": "CAL-3", "treatment": "降低", "goal": "SC-05: 侧信道防护",
             "requirement": "实施恒定时间算法和噪声屏蔽", "source": "M-005"},
            
            # TBox threats
            {"asset_id": "E001", "asset_name": "TBox", "subdomain1": "车载通信", "subdomain2": "远程通信", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Authenticity\n真实性", "stride": "S欺骗",
             "threat": "攻击者伪造云端服务器，执行中间人攻击",
             "attack_path": "1.DNS劫持\n2.证书伪造\n3.截获并篡改远程控制指令",
             "wp29_ref": "4.3.1", "vector": "网络", "complexity": "中", "privilege": "无", "interaction": "不需要",
             "safety": "S3严重", "safety_note": "可远程控制车辆", "safety_val": "3",
             "financial": "F3严重", "financial_note": "车辆被盗风险", "financial_val": "3",
             "operational": "O3严重", "operational_note": "远程功能被劫持", "operational_val": "3",
             "privacy": "P2中", "privacy_note": "通信内容泄露", "privacy_val": "2",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-06: 双向TLS认证",
             "requirement": "实施mTLS和证书固定", "source": "M-006"},
            
            # WiFi threats
            {"asset_id": "E002", "asset_name": "WiFi模块", "subdomain1": "车载通信", "subdomain2": "无线通信", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Authenticity\n真实性", "stride": "S欺骗",
             "threat": "恶意WiFi热点欺骗，诱骗车辆连接",
             "attack_path": "1.设置同名恶意AP\n2.车辆自动连接\n3.流量劫持窃取数据",
             "wp29_ref": "4.3.1", "vector": "邻居", "complexity": "低", "privilege": "无", "interaction": "需要",
             "safety": "S0无", "safety_note": "不直接影响安全", "safety_val": "0",
             "financial": "F2中", "financial_note": "数据窃取损失", "financial_val": "2",
             "operational": "O1低", "operational_note": "连接中断", "operational_val": "1",
             "privacy": "P3严重", "privacy_note": "通信被窃听", "privacy_val": "3",
             "risk_level": "CAL-3", "treatment": "降低", "goal": "SC-07: WiFi安全加固",
             "requirement": "禁用自动连接未知热点，强制WPA3", "source": "M-007"},
            
            # Bluetooth threats
            {"asset_id": "E003", "asset_name": "蓝牙模块", "subdomain1": "车载通信", "subdomain2": "无线通信", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Authenticity\n真实性", "stride": "S欺骗",
             "threat": "蓝牙身份欺骗，伪造已配对设备",
             "attack_path": "1.蓝牙嗅探获取配对信息\n2.克隆MAC地址\n3.伪装成已配对手机",
             "wp29_ref": "4.3.1", "vector": "邻居", "complexity": "中", "privilege": "无", "interaction": "不需要",
             "safety": "S0无", "safety_note": "不直接影响安全", "safety_val": "0",
             "financial": "F1低", "financial_note": "通话可能被窃听", "financial_val": "1",
             "operational": "O1低", "operational_note": "服务可能中断", "operational_val": "1",
             "privacy": "P2中", "privacy_note": "通讯录可能泄露", "privacy_val": "2",
             "risk_level": "CAL-2", "treatment": "降低", "goal": "SC-08: 蓝牙安全配对",
             "requirement": "使用Secure Simple Pairing", "source": "M-008"},
            
            # CAN Gateway threats
            {"asset_id": "E007", "asset_name": "CAN网关", "subdomain1": "车载网络", "subdomain2": "总线网关", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Integrity\n完整性", "stride": "T篡改",
             "threat": "CAN报文注入攻击，伪造车控指令",
             "attack_path": "1.通过OBD接口接入CAN总线\n2.逆向CAN报文协议\n3.注入伪造的控制指令",
             "wp29_ref": "5.1.1", "vector": "本地", "complexity": "中", "privilege": "无", "interaction": "不需要",
             "safety": "S3严重", "safety_note": "可能影响制动转向", "safety_val": "3",
             "financial": "F2中", "financial_note": "车辆损坏风险", "financial_val": "2",
             "operational": "O3严重", "operational_note": "车辆功能被劫持", "operational_val": "3",
             "privacy": "P0无", "privacy_note": "不涉及隐私", "privacy_val": "0",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-09: SecOC安全通信",
             "requirement": "实施AUTOSAR SecOC报文认证", "source": "M-009"},
            
            {"asset_id": "E007", "asset_name": "CAN网关", "subdomain1": "车载网络", "subdomain2": "总线网关", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Availability\n可用性", "stride": "D拒绝服务",
             "threat": "CAN总线洪泛攻击，导致通信瘫痪",
             "attack_path": "1.连接恶意设备到CAN总线\n2.发送高优先级报文洪泛\n3.总线过载通信中断",
             "wp29_ref": "4.3.4", "vector": "本地", "complexity": "低", "privilege": "无", "interaction": "不需要",
             "safety": "S3严重", "safety_note": "安全功能可能失效", "safety_val": "3",
             "financial": "F1低", "financial_note": "临时故障", "financial_val": "1",
             "operational": "O3严重", "operational_note": "车辆功能瘫痪", "operational_val": "3",
             "privacy": "P0无", "privacy_note": "不涉及隐私", "privacy_val": "0",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-10: 报文速率限制",
             "requirement": "实施网关报文过滤和速率限制", "source": "M-010"},
            
            # OBD threats
            {"asset_id": "E008", "asset_name": "OBD诊断接口", "subdomain1": "诊断接口", "subdomain2": "物理接口", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Authorization\n权限", "stride": "E提权",
             "threat": "诊断认证绕过，获取高权限诊断功能",
             "attack_path": "1.暴力破解种子密钥\n2.利用认证协议漏洞\n3.获取固件刷写权限",
             "wp29_ref": "4.3.5", "vector": "物理", "complexity": "中", "privilege": "无", "interaction": "不需要",
             "safety": "S2中", "safety_note": "固件可能被篡改", "safety_val": "2",
             "financial": "F2中", "financial_note": "车辆被非法改装", "financial_val": "2",
             "operational": "O2中", "operational_note": "参数可能被修改", "operational_val": "2",
             "privacy": "P1低", "privacy_note": "诊断数据泄露", "privacy_val": "1",
             "risk_level": "CAL-3", "treatment": "降低", "goal": "SC-11: 诊断安全访问",
             "requirement": "使用HSM实施安全访问服务", "source": "M-011"},
            
            # GNSS threats
            {"asset_id": "E010", "asset_name": "GNSS模块", "subdomain1": "定位系统", "subdomain2": "卫星通信", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Integrity\n完整性", "stride": "S欺骗",
             "threat": "GPS欺骗攻击，伪造位置信息",
             "attack_path": "1.使用GPS欺骗设备\n2.发送伪造卫星信号\n3.导航系统显示错误位置",
             "wp29_ref": "4.3.1", "vector": "邻居", "complexity": "中", "privilege": "无", "interaction": "不需要",
             "safety": "S1低", "safety_note": "可能导致迷路", "safety_val": "1",
             "financial": "F1低", "financial_note": "导航服务异常", "financial_val": "1",
             "operational": "O2中", "operational_note": "位置服务不准确", "operational_val": "2",
             "privacy": "P1低", "privacy_note": "位置追踪风险", "privacy_val": "1",
             "risk_level": "CAL-2", "treatment": "降低", "goal": "SC-12: GNSS信号验证",
             "requirement": "多源定位数据融合验证", "source": "M-012"},
            
            # V2X threats
            {"asset_id": "E011", "asset_name": "V2X模块", "subdomain1": "车路协同", "subdomain2": "无线通信", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Authenticity\n真实性", "stride": "S欺骗",
             "threat": "V2X消息伪造，发送虚假碰撞预警",
             "attack_path": "1.伪造V2X发射设备\n2.发送虚假BSM消息\n3.ADAS系统错误响应",
             "wp29_ref": "4.3.1", "vector": "邻居", "complexity": "中", "privilege": "无", "interaction": "不需要",
             "safety": "S3严重", "safety_note": "可能导致紧急制动", "safety_val": "3",
             "financial": "F1低", "financial_note": "可能造成交通事故", "financial_val": "1",
             "operational": "O2中", "operational_note": "驾驶辅助干扰", "operational_val": "2",
             "privacy": "P0无", "privacy_note": "不涉及隐私", "privacy_val": "0",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-13: V2X消息签名",
             "requirement": "实施SCMS PKI消息签名验证", "source": "M-013"},
            
            # UWB Digital Key threats
            {"asset_id": "E012", "asset_name": "UWB模块", "subdomain1": "数字钥匙", "subdomain2": "近场通信", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Authenticity\n真实性", "stride": "S欺骗",
             "threat": "数字钥匙中继攻击，远程解锁车辆",
             "attack_path": "1.中继设备A靠近车主\n2.中继设备B靠近车辆\n3.延长钥匙信号范围解锁",
             "wp29_ref": "4.3.1", "vector": "邻居", "complexity": "中", "privilege": "无", "interaction": "不需要",
             "safety": "S0无", "safety_note": "不直接影响安全", "safety_val": "0",
             "financial": "F3严重", "financial_note": "车辆被盗风险", "financial_val": "3",
             "operational": "O2中", "operational_note": "非法进入车辆", "operational_val": "2",
             "privacy": "P1低", "privacy_note": "车内物品被盗", "privacy_val": "1",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-14: UWB防中继",
             "requirement": "利用UWB测距检测中继攻击", "source": "M-014"},
            
            # DMS Camera threats
            {"asset_id": "E014", "asset_name": "DMS摄像头", "subdomain1": "驾驶监控", "subdomain2": "视觉传感", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Confidentiality\n机密性", "stride": "I信息泄露",
             "threat": "车内摄像头视频流被远程窃取",
             "attack_path": "1.云端服务漏洞利用\n2.远程访问摄像头\n3.实时窃取车内视频",
             "wp29_ref": "4.3.3", "vector": "网络", "complexity": "中", "privilege": "低", "interaction": "不需要",
             "safety": "S0无", "safety_note": "不影响安全", "safety_val": "0",
             "financial": "F2中", "financial_note": "隐私诉讼风险", "financial_val": "2",
             "operational": "O0无", "operational_note": "不影响操作", "operational_val": "0",
             "privacy": "P3严重", "privacy_note": "严重隐私侵犯", "privacy_val": "3",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-15: 摄像头访问控制",
             "requirement": "运行时权限管理，物理指示灯", "source": "M-015"},
            
            # Microphone threats
            {"asset_id": "E016", "asset_name": "麦克风阵列", "subdomain1": "语音输入", "subdomain2": "音频传感", "subdomain3": "N/A",
             "category": "外部实体", "sec_attr": "Confidentiality\n机密性", "stride": "I信息泄露",
             "threat": "车内语音被恶意应用窃听录制",
             "attack_path": "1.恶意应用获取麦克风权限\n2.后台静默录音\n3.上传至攻击者服务器",
             "wp29_ref": "4.3.3", "vector": "本地", "complexity": "低", "privilege": "低", "interaction": "需要",
             "safety": "S0无", "safety_note": "不影响安全", "safety_val": "0",
             "financial": "F1低", "financial_note": "隐私泄露风险", "financial_val": "1",
             "operational": "O0无", "operational_note": "不影响操作", "operational_val": "0",
             "privacy": "P3严重", "privacy_note": "私密对话泄露", "privacy_val": "3",
             "risk_level": "CAL-3", "treatment": "降低", "goal": "SC-16: 麦克风权限控制",
             "requirement": "严格运行时权限，录音指示", "source": "M-016"},
            
            # OTA threats
            {"asset_id": "P017", "asset_name": "OTA升级客户端", "subdomain1": "远程升级", "subdomain2": "固件更新", "subdomain3": "N/A",
             "category": "固件及应用", "sec_attr": "Integrity\n完整性", "stride": "T篡改",
             "threat": "OTA升级包被篡改，植入恶意代码",
             "attack_path": "1.中间人攻击截获升级包\n2.篡改固件内容\n3.车辆安装恶意固件",
             "wp29_ref": "5.1.1", "vector": "网络", "complexity": "高", "privilege": "无", "interaction": "不需要",
             "safety": "S3严重", "safety_note": "系统可能被完全控制", "safety_val": "3",
             "financial": "F3严重", "financial_note": "大规模召回风险", "financial_val": "3",
             "operational": "O3严重", "operational_note": "所有功能被劫持", "operational_val": "3",
             "privacy": "P2中", "privacy_note": "数据可能被窃取", "privacy_val": "2",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-17: OTA签名验证",
             "requirement": "RSA-4096签名验证，防回滚保护", "source": "M-017"},
            
            # Remote control threats
            {"asset_id": "P018", "asset_name": "远程车控服务", "subdomain1": "远程服务", "subdomain2": "车辆控制", "subdomain3": "N/A",
             "category": "固件及应用", "sec_attr": "Authenticity\n真实性", "stride": "S欺骗",
             "threat": "远程车控指令被伪造，非法控制车辆",
             "attack_path": "1.窃取用户凭证\n2.伪造远程控制请求\n3.非法解锁启动车辆",
             "wp29_ref": "4.3.1", "vector": "网络", "complexity": "中", "privilege": "低", "interaction": "不需要",
             "safety": "S1低", "safety_note": "可能被远程启动", "safety_val": "1",
             "financial": "F3严重", "financial_note": "车辆被盗风险", "financial_val": "3",
             "operational": "O3严重", "operational_note": "车辆被非法控制", "operational_val": "3",
             "privacy": "P1低", "privacy_note": "位置信息泄露", "privacy_val": "1",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-18: 多因素认证",
             "requirement": "远程控制需多因素认证确认", "source": "M-018"},
            
            # App Store threats
            {"asset_id": "P019", "asset_name": "应用商店", "subdomain1": "应用服务", "subdomain2": "软件分发", "subdomain3": "N/A",
             "category": "固件及应用", "sec_attr": "Integrity\n完整性", "stride": "T篡改",
             "threat": "恶意应用通过应用商店分发",
             "attack_path": "1.开发者上传恶意应用\n2.绕过审核机制\n3.用户安装后执行恶意代码",
             "wp29_ref": "4.3.2", "vector": "网络", "complexity": "中", "privilege": "低", "interaction": "需要",
             "safety": "S1低", "safety_note": "可能影响驾驶注意力", "safety_val": "1",
             "financial": "F2中", "financial_note": "数据被窃取", "financial_val": "2",
             "operational": "O2中", "operational_note": "系统资源被滥用", "operational_val": "2",
             "privacy": "P3严重", "privacy_note": "用户数据泄露", "privacy_val": "3",
             "risk_level": "CAL-4", "treatment": "降低", "goal": "SC-19: 应用签名验证",
             "requirement": "强制应用签名验证，安全审核", "source": "M-019"},
            
            # User data threats
            {"asset_id": "S006", "asset_name": "用户配置", "subdomain1": "数据存储", "subdomain2": "用户数据", "subdomain3": "N/A",
             "category": "数据存储", "sec_attr": "Confidentiality\n机密性", "stride": "I信息泄露",
             "threat": "用户账户凭证被窃取",
             "attack_path": "1.钓鱼攻击获取凭证\n2.本地存储漏洞\n3.账户被接管",
             "wp29_ref": "4.3.3", "vector": "网络", "complexity": "低", "privilege": "无", "interaction": "需要",
             "safety": "S0无", "safety_note": "不影响安全", "safety_val": "0",
             "financial": "F2中", "financial_note": "账户被滥用", "financial_val": "2",
             "operational": "O1低", "operational_note": "需要重置账户", "operational_val": "1",
             "privacy": "P3严重", "privacy_note": "个人信息泄露", "privacy_val": "3",
             "risk_level": "CAL-3", "treatment": "降低", "goal": "SC-20: 凭证安全存储",
             "requirement": "PBKDF2加密存储，异常登录检测", "source": "M-020"},
        ]

    def _apply_cell_style(self, cell, font=None, fill=None, alignment=None, border=True):
        """Apply styles to a cell."""
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = self.border

    def _merge_and_style(self, ws, start_row, start_col, end_row, end_col, value, font=None, fill=None, alignment=None):
        """Merge cells and apply style."""
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col, value=value)
        self._apply_cell_style(cell, font, fill, alignment or self.center_align)
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = self.border

    def generate(self, template_file, output_file):
        """Generate expanded report based on template."""
        # Load template
        wb = load_workbook(template_file)
        
        # Update sheets
        self._update_asset_list(wb)
        self._update_data_flow(wb)
        self._update_attack_tree(wb)
        self._update_tara_results(wb)
        
        # Save
        wb.save(output_file)
        print(f"✓ 扩展报告已生成: {output_file}")
        print(f"  - 资产数量: {len(self.assets)}")
        print(f"  - 威胁数量: {len(self.threats)}")

    def _update_asset_list(self, wb):
        """Update asset list sheet (Tab 2)."""
        ws = wb["2-资产列表"]
        
        # Clear existing data (keep headers)
        for row in range(5, ws.max_row + 50):
            for col in range(1, 14):
                ws.cell(row=row, column=col).value = None
        
        # Write new asset data
        row = 5
        for asset in self.assets:
            ws.cell(row=row, column=1, value=asset["id"])
            ws.cell(row=row, column=2, value=asset["name"])
            ws.cell(row=row, column=3, value=asset["category"])
            ws.cell(row=row, column=4, value=asset["desc"])
            ws.cell(row=row, column=5, value=asset["auth"])
            ws.cell(row=row, column=6, value=asset["integ"])
            ws.cell(row=row, column=7, value=asset["nonrep"])
            ws.cell(row=row, column=8, value=asset["conf"])
            ws.cell(row=row, column=9, value=asset["avail"])
            ws.cell(row=row, column=10, value=asset["authz"])
            
            # Apply styles
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                fill = self.alt_fill if row % 2 == 0 else None
                self._apply_cell_style(cell, self.data_font, fill, 
                                      self.center_align if col >= 5 else self.left_align)
            row += 1

    def _update_data_flow(self, wb):
        """Update data flow diagram sheet (Tab 3)."""
        ws = wb["3-数据流图"]
        
        # Unmerge all merged cells first
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
        
        # Create new data flow diagram
        data_flow_diagram = """
┌──────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────┐
│                                    MY25 EV平台 IVI系统数据流图 Data Flow Diagram (Level 0)                                │
│                                                                                                                          │
│     ┌─────────────┐                                                                      ┌─────────────┐                 │
│     │ 云服务器    │                                                                      │  车身控制器  │                 │
│     │ Cloud Server│                                                                      │  BCM/Gateway│                 │
│     │    (E1)     │                                                                      │    (E7)     │                 │
│     └──────┬──────┘                                                                      └──────┬──────┘                 │
│            │                                                                                    │                         │
│            │ DF1: OTA升级包/远程控制                                                            │ DF3: CAN车控指令         │
│            ↓                                                                                    ↑                         │
│     ┌─────────────┐                                ╔══════════════════════════════════════════════════════════════╗      │
│     │   TBox      │ ────────DF2: 远程指令──────→  ║                                                              ║      │
│     │ 5G Modem    │ ←───────DF4: 状态上报──────── ║            IVI 主系统 (信任边界)                             ║      │
│     │    (E1)     │                                ║            IVI Main System                                   ║      │
│     └─────────────┘                                ║                                                              ║      │
│                                                    ║  ┌────────────────────────────────────────────────────────┐  ║      │
│     ┌─────────────┐                                ║  │                P1: SOC 主处理器                        │  ║      │
│     │  WiFi模块   │ ────────DF5: 网络数据──────→  ║  │  ┌──────────┐ ┌──────────┐ ┌──────────┐ ┌──────────┐  │  ║──────→│
│     │ WiFi 6E     │                                ║  │  │ S1:系统  │ │ S2:用户  │ │ S3:密钥  │ │ S4:导航  │  │  ║      │
│     │    (E2)     │                                ║  │  │   存储   │ │   数据   │ │   存储   │ │   历史   │  │  ║      │
│     └─────────────┘                                ║  │  └──────────┘ └──────────┘ └──────────┘ └──────────┘  │  ║      │
│                                                    ║  │                        ↕                               │  ║      │
│     ┌─────────────┐                                ║  │         ┌────────────────────────────────┐             │  ║      │
│     │  蓝牙模块   │ ────────DF6: 音频/通讯录───→  ║  │         │      P2: MCU 安全控制器         │             │  ║      │
│     │ Bluetooth   │ ←───────DF7: 媒体控制──────── ║  │         │   ┌──────┐  ┌──────┐           │             │  ║      │
│     │    (E3)     │                                ║  │         │   │ HSM  │  │  SE  │           │             │  ║      │
│     └─────────────┘                                ║  │         │   │(P5)  │  │ (P6) │           │             │  ║      │
│                                                    ║  │         │   └──────┘  └──────┘           │             │  ║      │
│     ┌─────────────┐                                ║  │         └────────────────────────────────┘             │  ║      │
│     │  用户手机   │ ────────DF8: CarPlay数据───→  ║  └────────────────────────────────────────────────────────┘  ║      │
│     │ User Phone  │ ←───────DF9: 显示投屏──────── ║                           ↕                                  ║      │
│     │    (E4)     │                                ║  ┌────────────────────────────────────────────────────────┐  ║      │
│     └─────────────┘                                ║  │  E14: DMS   E16: 麦克风   E17-19: 显示屏                │  ║      │
│                                                    ║  │  摄像头      阵列          (中控/仪表/HUD)              │  ║      │
│     ┌─────────────┐                                ║  └────────────────────────────────────────────────────────┘  ║      │
│     │  诊断设备   │ ────────DF10: 诊断请求─────→  ╚══════════════════════════════════════════════════════════════╝      │
│     │ Diagnostic  │ ←───────DF11: 诊断响应─────── │                                                                      │
│     │ Tool (E8)   │                                │                                                                      │
│     └─────────────┘                                │                                                                      │
│                                                    │         ┌─────────────┐      ┌─────────────┐                        │
│     ┌─────────────┐                                │         │  V2X模块    │      │  UWB/NFC    │                        │
│     │  USB设备    │ ────────DF12: 媒体文件─────→  │         │    (E11)    │      │  (E12/E13)  │                        │
│     │ USB Device  │                                │         └──────┬──────┘      └──────┬──────┘                        │
│     │    (E6)     │                                │                │ DF13             │ DF14                             │
│     └─────────────┘                                │                ↓ V2X消息          ↓ 数字钥匙                          │
│                                                    │         路侧单元/其他车辆      用户手机/卡片                          │
│                                                                                                                          │
│  图例:  ═══ 信任边界 Trust Boundary    ───→ 数据流 Data Flow    [Px] 处理 Process    [Ex] 外部实体    [Sx] 数据存储        │
└──────────────────────────────────────────────────────────────────────────────────────────────────────────────────────────┘
"""
        
        # Clear old content and write new
        for row in range(3, 42):
            ws.cell(row=row, column=1).value = None
            
        ws.cell(row=3, column=1, value=data_flow_diagram)
        ws.cell(row=3, column=1).font = Font(name='Courier New', size=9)
        ws.cell(row=3, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=3, start_column=1, end_row=41, end_column=9)
        
        # Update data flow list
        data_flows = [
            ("DF-01", "OTA升级包", "云服务器→TBox→IVI", "整车OTA固件更新"),
            ("DF-02", "远程控制指令", "云服务器→TBox→MCU", "手机APP远程车控"),
            ("DF-03", "CAN车控报文", "MCU↔CAN网关↔BCM", "车身控制指令"),
            ("DF-04", "车辆状态上报", "IVI→TBox→云服务器", "车辆状态监控"),
            ("DF-05", "WiFi网络数据", "WiFi模块↔SOC", "互联网接入"),
            ("DF-06", "蓝牙音频/通讯录", "手机→蓝牙→IVI", "多媒体和通讯"),
            ("DF-07", "媒体控制", "IVI→蓝牙→手机", "AVRCP控制"),
            ("DF-08", "CarPlay/AA数据", "手机→USB/WiFi→IVI", "手机投屏"),
            ("DF-09", "显示投屏", "SOC→显示屏", "多屏显示输出"),
            ("DF-10", "诊断请求", "诊断工具→OBD→MCU", "UDS诊断请求"),
            ("DF-11", "诊断响应", "MCU→OBD→诊断工具", "UDS诊断响应"),
            ("DF-12", "媒体文件", "USB设备→SOC", "U盘媒体导入"),
            ("DF-13", "V2X消息", "RSU/OBU↔V2X模块", "车路协同通信"),
            ("DF-14", "数字钥匙信号", "手机→UWB/NFC→SE", "无感解锁/启动"),
        ]
        
        # Write data flow list
        row = 43
        ws.cell(row=row, column=1, value="数据流清单 Data Flow List")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1
        
        headers = ["数据流ID\nDF ID", "名称\nName", "路径\nPath", "描述\nDescription"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_cell_style(cell, self.header_font, self.header_fill, self.center_align)
        row += 1
        
        for df in data_flows:
            for col, value in enumerate(df, 1):
                cell = ws.cell(row=row, column=col, value=value)
                fill = self.alt_fill if row % 2 == 0 else None
                self._apply_cell_style(cell, self.data_font, fill, self.left_align)
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 100
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30

    def _update_attack_tree(self, wb):
        """Update attack tree sheet (Tab 4)."""
        ws = wb["4-攻击树图"]
        
        # Unmerge all merged cells first
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
        
        # Clear old content
        for row in range(1, 100):
            ws.cell(row=row, column=1).value = None
        
        # Title
        ws.cell(row=1, column=1, value="MY25 EV平台中控主机 - 攻击树分析 Attack Tree Analysis")
        ws.cell(row=1, column=1).font = self.title_font
        ws.cell(row=1, column=1).fill = self.title_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        # Attack Tree 1: Remote Attack
        attack_tree1 = """
                                        ┌─────────────────────────────────┐
                                        │     【攻击目标 Attack Goal】     │
                                        │     远程入侵IVI系统              │
                                        │     Remote Compromise IVI       │
                                        └────────────────┬────────────────┘
                                                         │
                     ┌───────────────────────────────────┼───────────────────────────────────┐
                     │                                   │                                   │
         ┌───────────┴───────────┐           ┌───────────┴───────────┐           ┌───────────┴───────────┐
         │ [OR] 通过TBox入侵     │           │ [OR] 通过WiFi入侵     │           │ [OR] 通过云服务入侵   │
         │ Via TBox (AT1.1)      │           │ Via WiFi (AT1.2)      │           │ Via Cloud (AT1.3)     │
         └───────────┬───────────┘           └───────────┬───────────┘           └───────────┬───────────┘
                     │                                   │                                   │
     ┌───────────────┼───────────────┐       ┌───────────┼───────────┐       ┌───────────────┼───────────────┐
     │               │               │       │           │           │       │               │               │
┌────┴────┐    ┌─────┴─────┐   ┌─────┴─────┐ │     ┌─────┴─────┐     │ ┌─────┴─────┐   ┌─────┴─────┐   ┌─────┴─────┐
│ 利用5G  │    │ 伪造远程  │   │ TBox固件  │ │     │ WiFi协议  │     │ │ 伪造OTA   │   │ 中间人    │   │ 云服务器  │
│ 协议漏洞│    │ 控制指令  │   │ 漏洞利用  │ │     │ 漏洞利用  │     │ │ 升级包    │   │ 攻击MITM  │   │ 被攻陷    │
│(1.1.1)  │    │ (1.1.2)   │   │ (1.1.3)   │ │     │ (1.2.1)   │     │ │ (1.3.1)   │   │ (1.3.2)   │   │ (1.3.3)   │
└─────────┘    └───────────┘   └───────────┘ │     └───────────┘     │ └───────────┘   └───────────┘   └───────────┘
                                             │           │           │
                                       ┌─────┴─────┐┌────┴────┐┌─────┴─────┐
                                       │ 恶意WiFi  ││连接恶意 ││ 无线监听  │
                                       │ AP欺骗   ││ 热点    ││ 数据窃取  │
                                       │ (1.2.2)  ││(1.2.3)  ││ (1.2.4)   │
                                       └──────────┘└─────────┘└───────────┘
"""
        ws.cell(row=3, column=1, value="攻击树1 Attack Tree 1: 远程入侵IVI系统 Remote Compromise of IVI System")
        ws.cell(row=3, column=1).font = Font(bold=True, size=11)
        ws.cell(row=4, column=1, value=attack_tree1)
        ws.cell(row=4, column=1).font = Font(name='Courier New', size=9)
        ws.cell(row=4, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=4, start_column=1, end_row=22, end_column=6)
        
        # Attack Tree 2: Physical Attack
        attack_tree2 = """
                                        ┌─────────────────────────────────┐
                                        │     【攻击目标 Attack Goal】     │
                                        │     物理接口攻击IVI系统          │
                                        │     Physical Attack on IVI      │
                                        └────────────────┬────────────────┘
                                                         │
                     ┌───────────────────────────────────┼───────────────────────────────────┐
                     │                                   │                                   │
         ┌───────────┴───────────┐           ┌───────────┴───────────┐           ┌───────────┴───────────┐
         │ [OR] OBD接口攻击      │           │ [OR] USB接口攻击      │           │ [OR] 调试接口攻击     │
         │ Via OBD (AT2.1)       │           │ Via USB (AT2.2)       │           │ Via Debug (AT2.3)     │
         └───────────┬───────────┘           └───────────┬───────────┘           └───────────┬───────────┘
                     │                                   │                                   │
     ┌───────────────┼───────────────┐       ┌───────────┼───────────┐       ┌───────────────┼───────────────┐
     │               │               │       │           │           │       │               │               │
┌────┴────┐    ┌─────┴─────┐   ┌─────┴─────┐┌────┴────┐┌────┴────┐┌────┴────┐┌────┴────┐   ┌─────┴─────┐   ┌─────┴─────┐
│诊断认证 │    │ CAN报文   │   │ 读取诊断  ││恶意U盘  ││固件刷写 ││USB漏洞  ││JTAG提取 │   │ UART命令  │   │ 故障注入  │
│绕过攻击 │    │ 注入伪造  │   │ 敏感数据  ││植入恶意 ││攻击     ││利用     ││密钥固件 │   │ 注入攻击  │   │ 绕过启动  │
│(2.1.1)  │    │ (2.1.2)   │   │ (2.1.3)   ││(2.2.1)  ││(2.2.2)  ││(2.2.3)  ││(2.3.1)  │   │ (2.3.2)   │   │ (2.3.3)   │
└─────────┘    └───────────┘   └───────────┘└─────────┘└─────────┘└─────────┘└─────────┘   └───────────┘   └───────────┘
"""
        ws.cell(row=24, column=1, value="攻击树2 Attack Tree 2: 物理接口攻击 Physical Interface Attack")
        ws.cell(row=24, column=1).font = Font(bold=True, size=11)
        ws.cell(row=25, column=1, value=attack_tree2)
        ws.cell(row=25, column=1).font = Font(name='Courier New', size=9)
        ws.cell(row=25, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=25, start_column=1, end_row=43, end_column=6)
        
        # Attack Tree 3: Data Theft
        attack_tree3 = """
                                        ┌─────────────────────────────────┐
                                        │     【攻击目标 Attack Goal】     │
                                        │     窃取用户隐私数据             │
                                        │     User Privacy Data Theft     │
                                        └────────────────┬────────────────┘
                                                         │
                     ┌───────────────────────────────────┼───────────────────────────────────┐
                     │                                   │                                   │
         ┌───────────┴───────────┐           ┌───────────┴───────────┐           ┌───────────┴───────────┐
         │ [OR] 车内监控窃取     │           │ [OR] 通信数据窃取     │           │ [OR] 存储数据提取     │
         │ In-cabin Spy (AT3.1) │           │ Comm Intercept(AT3.2) │           │ Storage Extract(AT3.3)│
         └───────────┬───────────┘           └───────────┬───────────┘           └───────────┬───────────┘
                     │                                   │                                   │
     ┌───────────────┼───────────────┐       ┌───────────┼───────────┐       ┌───────────────┼───────────────┐
     │               │               │       │           │           │       │               │               │
┌────┴────┐    ┌─────┴─────┐   ┌─────┴─────┐┌────┴────┐┌────┴────┐┌────┴────┐┌────┴────┐   ┌─────┴─────┐   ┌─────┴─────┐
│DMS摄像头│    │ OMS摄像头 │   │ 麦克风    ││WiFi流量 ││蓝牙通讯 ││V2X消息  ││eMMC存储 │   │ 用户配置  │   │ 导航轨迹  │
│视频窃取 │    │ 视频窃取  │   │ 语音窃听  ││窃听     ││录数据   ││位置追踪 ││固件提取 │   │ 数据窃取  │   │ 数据窃取  │
│(3.1.1)  │    │ (3.1.2)   │   │ (3.1.3)   ││(3.2.1)  ││(3.2.2)  ││(3.2.3)  ││(3.3.1)  │   │ (3.3.2)   │   │ (3.3.3)   │
└─────────┘    └───────────┘   └───────────┘└─────────┘└─────────┘└─────────┘└─────────┘   └───────────┘   └───────────┘
"""
        ws.cell(row=45, column=1, value="攻击树3 Attack Tree 3: 数据窃取 Data Theft")
        ws.cell(row=45, column=1).font = Font(bold=True, size=11)
        ws.cell(row=46, column=1, value=attack_tree3)
        ws.cell(row=46, column=1).font = Font(name='Courier New', size=9)
        ws.cell(row=46, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=46, start_column=1, end_row=64, end_column=6)
        
        # Attack Tree 4: Vehicle Control
        attack_tree4 = """
                                        ┌─────────────────────────────────┐
                                        │     【攻击目标 Attack Goal】     │
                                        │     非法控制车辆功能             │
                                        │     Unauthorized Vehicle Ctrl   │
                                        └────────────────┬────────────────┘
                                                         │
                     ┌───────────────────────────────────┼───────────────────────────────────┐
                     │                                   │                                   │
         ┌───────────┴───────────┐           ┌───────────┴───────────┐           ┌───────────┴───────────┐
         │ [OR] CAN总线攻击      │           │ [OR] 数字钥匙攻击     │           │ [OR] 远程控制劫持     │
         │ Via CAN (AT4.1)       │           │ Via DK (AT4.2)        │           │ Remote Hijack(AT4.3)  │
         └───────────┬───────────┘           └───────────┬───────────┘           └───────────┬───────────┘
                     │                                   │                                   │
     ┌───────────────┼───────────────┐       ┌───────────┼───────────┐       ┌───────────────┼───────────────┐
     │               │               │       │           │           │       │               │               │
┌────┴────┐    ┌─────┴─────┐   ┌─────┴─────┐┌────┴────┐┌────┴────┐┌────┴────┐┌────┴────┐   ┌─────┴─────┐   ┌─────┴─────┐
│CAN报文  │    │ CAN DoS   │   │ SecOC    ││UWB中继  ││NFC克隆  ││蓝牙钥匙 ││APP凭证  │   │ 服务器    │   │ 伪造远程  │
│注入伪造 │    │ 洪泛攻击  │   │ 密钥破解 ││攻击     ││攻击     ││欺骗     ││窃取     │   │ 入侵      │   │ 控制指令  │
│(4.1.1)  │    │ (4.1.2)   │   │ (4.1.3)  ││(4.2.1)  ││(4.2.2)  ││(4.2.3)  ││(4.3.1)  │   │ (4.3.2)   │   │ (4.3.3)   │
└─────────┘    └───────────┘   └───────────┘└─────────┘└─────────┘└─────────┘└─────────┘   └───────────┘   └───────────┘
"""
        ws.cell(row=66, column=1, value="攻击树4 Attack Tree 4: 恶意车辆控制 Malicious Vehicle Control")
        ws.cell(row=66, column=1).font = Font(bold=True, size=11)
        ws.cell(row=67, column=1, value=attack_tree4)
        ws.cell(row=67, column=1).font = Font(name='Courier New', size=9)
        ws.cell(row=67, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=67, start_column=1, end_row=85, end_column=6)
        
        ws.column_dimensions['A'].width = 120

    def _update_tara_results(self, wb):
        """Update TARA results sheet (Tab 5)."""
        ws = wb["5-TARA分析结果"]
        
        # Unmerge data cells (keep header merges)
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            # Only unmerge cells in data rows (row 6 and below)
            if merged_range.min_row >= 6:
                ws.unmerge_cells(str(merged_range))
        
        # Clear old data (keep headers in rows 3-5)
        for row in range(6, ws.max_row + 50):
            for col in range(1, 41):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, type(None)):
                    try:
                        cell.value = None
                    except AttributeError:
                        pass
        
        # Write threat data
        row = 6
        for threat in self.threats:
            # Asset identification
            ws.cell(row=row, column=1, value=threat["asset_id"])
            ws.cell(row=row, column=2, value=threat["asset_name"])
            ws.cell(row=row, column=3, value=threat["subdomain1"])
            ws.cell(row=row, column=4, value=threat["subdomain2"])
            ws.cell(row=row, column=5, value=threat["subdomain3"])
            ws.cell(row=row, column=6, value=threat["category"])
            
            # Threat scenario
            ws.cell(row=row, column=7, value=threat["sec_attr"])
            ws.cell(row=row, column=8, value=threat["stride"])
            ws.cell(row=row, column=9, value=threat["threat"])
            ws.cell(row=row, column=10, value=threat["attack_path"])
            ws.cell(row=row, column=11, value=threat["wp29_ref"])
            
            # Threat analysis
            ws.cell(row=row, column=12, value=threat["vector"])
            ws.cell(row=row, column=13, value="0.5")  # Value placeholder
            ws.cell(row=row, column=14, value=threat["complexity"])
            ws.cell(row=row, column=15, value="0.5")
            ws.cell(row=row, column=16, value=threat["privilege"])
            ws.cell(row=row, column=17, value="0.5")
            ws.cell(row=row, column=18, value=threat["interaction"])
            ws.cell(row=row, column=19, value="0.5")
            ws.cell(row=row, column=20, value="=L{0}*M{0}*N{0}*O{0}".format(row))  # Calc
            ws.cell(row=row, column=21, value="Medium")  # Level
            
            # Impact analysis
            ws.cell(row=row, column=22, value=threat["safety"])
            ws.cell(row=row, column=23, value=threat["safety_note"])
            ws.cell(row=row, column=24, value=threat["safety_val"])
            ws.cell(row=row, column=25, value=threat["financial"])
            ws.cell(row=row, column=26, value=threat["financial_note"])
            ws.cell(row=row, column=27, value=threat["financial_val"])
            ws.cell(row=row, column=28, value=threat["operational"])
            ws.cell(row=row, column=29, value=threat["operational_note"])
            ws.cell(row=row, column=30, value=threat["operational_val"])
            ws.cell(row=row, column=31, value=threat["privacy"])
            ws.cell(row=row, column=32, value=threat["privacy_note"])
            ws.cell(row=row, column=33, value=threat["privacy_val"])
            ws.cell(row=row, column=34, value="=MAX(X{0},AA{0},AD{0},AG{0})".format(row))  # Impact calc
            ws.cell(row=row, column=35, value="Major")  # Impact level
            
            # Risk assessment and treatment
            ws.cell(row=row, column=36, value=threat["risk_level"])
            ws.cell(row=row, column=37, value=threat["treatment"])
            ws.cell(row=row, column=38, value=threat["goal"])
            ws.cell(row=row, column=39, value=threat["requirement"])
            ws.cell(row=row, column=40, value=threat["source"])
            
            # Apply styles
            for col in range(1, 41):
                cell = ws.cell(row=row, column=col)
                fill = self.alt_fill if row % 2 == 0 else None
                
                # Color risk level
                if col == 36:
                    risk = threat["risk_level"]
                    if risk in self.risk_fills:
                        fill = self.risk_fills[risk]
                
                self._apply_cell_style(cell, self.data_font, fill, self.center_align)
            
            row += 1


def main():
    """Generate the expanded TARA report."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    template_file = os.path.join(script_dir, 'MY25 EV平台中控主机_TARA分析报告.xlsx')
    output_file = os.path.join(script_dir, 'MY25_EV平台中控主机_TARA分析报告_扩展版V2.xlsx')
    
    print("=" * 60)
    print("MY25 EV平台中控主机 TARA分析报告扩展生成")
    print("=" * 60)
    
    generator = TARAReportGenerator()
    generator.generate(template_file, output_file)
    
    print("\n" + "=" * 60)
    print("报告生成完成!")
    print("=" * 60)


if __name__ == "__main__":
    main()
