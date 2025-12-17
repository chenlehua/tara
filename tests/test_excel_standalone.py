"""
Standalone Test for Excel Generator Sample Data
================================================

This test verifies the sample data generation logic without
requiring the full application context.
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_sample_tara_data(project: dict) -> dict:
    """Generate sample TARA analysis data for demonstration purposes."""
    project_name = project.get("name", "车载信息娱乐系统")
    
    # Sample assets
    sample_assets = [
        {
            "id": 1, "name": "SOC主处理器", "asset_type": "ECU", "category": "内部实体",
            "description": "高通SA8155P系统级芯片",
            "security_attrs": {"authenticity": True, "integrity": True, "confidentiality": True, "availability": True},
            "interfaces": [{"type": "Ethernet"}, {"type": "USB"}]
        },
        {
            "id": 2, "name": "MCU安全控制器", "asset_type": "ECU", "category": "内部实体",
            "description": "NXP S32G安全微控制器",
            "security_attrs": {"authenticity": True, "integrity": True, "confidentiality": True, "availability": True},
            "interfaces": [{"type": "CAN"}, {"type": "Ethernet"}]
        },
        {
            "id": 3, "name": "WiFi/BT通信模块", "asset_type": "通信模块", "category": "外部接口",
            "description": "WiFi 6和蓝牙5.2模块",
            "security_attrs": {"authenticity": True, "integrity": True, "confidentiality": True, "availability": True},
            "interfaces": [{"type": "WiFi"}, {"type": "Bluetooth"}]
        },
        {
            "id": 4, "name": "4G/5G T-Box", "asset_type": "通信模块", "category": "外部接口",
            "description": "车载远程信息处理单元",
            "security_attrs": {"authenticity": True, "integrity": True, "confidentiality": True, "availability": True},
            "interfaces": [{"type": "4G/5G"}]
        },
        {
            "id": 5, "name": "CAN总线网关", "asset_type": "网关", "category": "内部接口",
            "description": "CAN FD网关",
            "security_attrs": {"authenticity": True, "integrity": True, "availability": True},
            "interfaces": [{"type": "CAN"}]
        },
    ]
    
    # Sample threats
    sample_threats = [
        {
            "id": 1, "asset_id": 4, "asset_name": "4G/5G T-Box",
            "threat_type": "S", "category": "Spoofing",
            "name": "远程服务器身份伪造攻击",
            "description": "攻击者伪造合法云端服务器",
            "attack_vector": "Network远程网络攻击",
            "attack_path": "4G/5G→T-Box→网关",
            "wp29_ref": "4.3.1",
            "safety_impact": 3, "financial_impact": 3, "operational_impact": 3, "privacy_impact": 2,
            "impact_level": 3, "risk_level": "CAL-4", "iso_clause": "9.4"
        },
        {
            "id": 2, "asset_id": 5, "asset_name": "CAN总线网关",
            "threat_type": "T", "category": "Tampering",
            "name": "CAN报文注入攻击",
            "description": "向CAN总线注入恶意报文",
            "attack_vector": "Local本地访问",
            "attack_path": "OBD-II→CAN网关",
            "wp29_ref": "5.1.1",
            "safety_impact": 4, "financial_impact": 3, "operational_impact": 4, "privacy_impact": 0,
            "impact_level": 4, "risk_level": "CAL-4", "iso_clause": "9.5"
        },
        {
            "id": 3, "asset_id": 3, "asset_name": "WiFi/BT通信模块",
            "threat_type": "I", "category": "Information Disclosure",
            "name": "用户隐私数据泄露",
            "description": "窃取用户敏感信息",
            "attack_vector": "Network远程网络",
            "attack_path": "WiFi→数据存储",
            "wp29_ref": "4.3.3",
            "safety_impact": 0, "financial_impact": 3, "operational_impact": 1, "privacy_impact": 4,
            "impact_level": 4, "risk_level": "CAL-3", "iso_clause": "9.6"
        },
        {
            "id": 4, "asset_id": 5, "asset_name": "CAN总线网关",
            "threat_type": "D", "category": "Denial of Service",
            "name": "CAN总线洪泛攻击",
            "description": "导致总线瘫痪",
            "attack_vector": "Local本地访问",
            "attack_path": "恶意ECU→CAN总线",
            "wp29_ref": "4.3.4",
            "safety_impact": 4, "financial_impact": 2, "operational_impact": 4, "privacy_impact": 0,
            "impact_level": 4, "risk_level": "CAL-4", "iso_clause": "9.7"
        },
        {
            "id": 5, "asset_id": 1, "asset_name": "SOC主处理器",
            "threat_type": "E", "category": "Elevation of Privilege",
            "name": "Android系统提权攻击",
            "description": "获取Root权限",
            "attack_vector": "Network远程网络",
            "attack_path": "恶意应用→系统漏洞",
            "wp29_ref": "4.3.5",
            "safety_impact": 4, "financial_impact": 4, "operational_impact": 4, "privacy_impact": 4,
            "impact_level": 4, "risk_level": "CAL-4", "iso_clause": "9.8"
        },
    ]
    
    # Sample control measures
    sample_measures = [
        {"id": 1, "threat_id": 1, "threat_risk_id": 1, "name": "双向TLS认证", 
         "implementation": "部署mTLS确保服务器和客户端双向认证", "iso21434_ref": "RQ-09-01"},
        {"id": 2, "threat_id": 2, "threat_risk_id": 2, "name": "消息认证码(MAC)", 
         "implementation": "对CAN报文添加AES-128-CMAC认证", "iso21434_ref": "RQ-09-04"},
        {"id": 3, "threat_id": 2, "threat_risk_id": 2, "name": "入侵检测系统", 
         "implementation": "部署CAN IDS检测异常报文", "iso21434_ref": "RQ-09-05"},
        {"id": 4, "threat_id": 3, "threat_risk_id": 3, "name": "数据加密存储", 
         "implementation": "使用AES-256加密敏感数据", "iso21434_ref": "RQ-09-08"},
        {"id": 5, "threat_id": 4, "threat_risk_id": 4, "name": "报文速率限制", 
         "implementation": "实施CAN报文速率限制", "iso21434_ref": "RQ-09-12"},
        {"id": 6, "threat_id": 5, "threat_risk_id": 5, "name": "SELinux强制访问控制", 
         "implementation": "启用SELinux限制进程权限", "iso21434_ref": "RQ-09-15"},
    ]
    
    # Risk distribution
    risk_distribution = {"CAL-4": 0, "CAL-3": 0, "CAL-2": 0, "CAL-1": 0}
    for threat in sample_threats:
        risk_level = threat.get("risk_level", "CAL-2")
        if risk_level in risk_distribution:
            risk_distribution[risk_level] += 1
    
    return {
        "assets": sample_assets,
        "threats": sample_threats,
        "control_measures": sample_measures,
        "risk_distribution": risk_distribution,
    }


def create_tara_results_sheet(wb, assets, threats, measures, risk_dist):
    """Create the TARA Analysis Results sheet."""
    ws = wb.create_sheet("5-TARA分析结果")
    
    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    subheader_font = Font(bold=True, size=9)
    subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    data_font = Font(size=9)
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    risk_fills = {
        'CAL-4': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        'CAL-3': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        'CAL-2': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        'CAL-1': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    }
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    cell = ws.cell(row=1, column=1, value="TARA分析结果 TARA Analysis Results")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center_align
    cell.border = border
    
    # Headers
    headers = [
        "资产ID", "资产名称", "分类", "STRIDE", "威胁名称", "威胁描述", "攻击向量",
        "攻击路径", "安全影响", "经济影响", "操作影响", "隐私影响", "影响等级",
        "风险等级", "风险处置", "安全目标", "安全需求", "WP29 Ref", "ISO Ref", "状态"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Build measures lookup
    measures_by_threat = {}
    for m in measures:
        tid = m.get("threat_id") or m.get("threat_risk_id")
        if tid:
            if tid not in measures_by_threat:
                measures_by_threat[tid] = []
            measures_by_threat[tid].append(m)
    
    # Data rows
    row = 4
    for threat in threats:
        asset_id = threat.get("asset_id", "")
        asset = next((a for a in assets if a.get("id") == asset_id), {})
        
        stride_map = {
            'S': "S-欺骗", 'T': "T-篡改", 'R': "R-抵赖",
            'I': "I-泄露", 'D': "D-DoS", 'E': "E-提权"
        }
        threat_type = threat.get("threat_type", "T")
        
        # Get measures for this threat
        threat_measures = measures_by_threat.get(threat.get("id"), [])
        sec_goal = threat_measures[0].get("name", "-") if threat_measures else "-"
        sec_req = threat_measures[0].get("implementation", "-")[:30] if threat_measures else "-"
        
        risk_level = threat.get("risk_level", "CAL-2")
        treatment_map = {
            "CAL-4": "降低", "CAL-3": "降低", "CAL-2": "降低/接受", "CAL-1": "接受"
        }
        
        row_data = [
            f"A-{asset_id:03d}" if isinstance(asset_id, int) else str(asset_id),
            asset.get("name", threat.get("asset_name", ""))[:15],
            asset.get("category", "内部实体"),
            stride_map.get(threat_type, "T-篡改"),
            threat.get("name", threat.get("threat_name", ""))[:20],
            threat.get("description", "")[:25],
            threat.get("attack_vector", "-")[:15],
            threat.get("attack_path", "-")[:15],
            f"S{threat.get('safety_impact', 0)}",
            f"F{threat.get('financial_impact', 0)}",
            f"O{threat.get('operational_impact', 0)}",
            f"P{threat.get('privacy_impact', 0)}",
            threat.get("impact_level", 0),
            risk_level,
            treatment_map.get(risk_level, "降低"),
            sec_goal[:15],
            sec_req[:20],
            threat.get("wp29_ref", "-"),
            threat.get("iso_clause", "-"),
            "已识别"
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align if col >= 4 else left_align
            cell.border = border
            
            fill = alt_fill if row % 2 == 0 else None
            if col == 14 and value in risk_fills:  # Risk level column
                fill = risk_fills[value]
            if fill:
                cell.fill = fill
        
        row += 1
    
    # Set column widths
    col_widths = [8, 15, 10, 10, 20, 25, 15, 15, 8, 8, 8, 8, 10, 10, 12, 15, 20, 8, 8, 8]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    return ws


def generate_excel(project):
    """Generate a complete Excel report."""
    wb = Workbook()
    
    # Generate sample data
    sample_data = generate_sample_tara_data(project)
    
    # Create TARA results sheet
    create_tara_results_sheet(
        wb,
        sample_data["assets"],
        sample_data["threats"],
        sample_data["control_measures"],
        sample_data["risk_distribution"]
    )
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def main():
    """Test the Excel generation."""
    print("=" * 60)
    print("Testing TARA Excel Report Generation with Sample Data")
    print("=" * 60)
    
    project = {
        "name": "车载IVI系统TARA分析",
        "description": "高端乘用车信息娱乐系统安全分析",
        "vehicle_type": "高端乘用车",
    }
    
    # Generate Excel
    buffer = generate_excel(project)
    
    # Check results
    file_size = buffer.getbuffer().nbytes
    print(f"\n✓ Excel file generated successfully!")
    print(f"  File size: {file_size:,} bytes")
    
    # Save for inspection
    output_path = "tests/output/tara_sample_report.xlsx"
    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "wb") as f:
        f.write(buffer.getvalue())
    print(f"  Output file: {output_path}")
    
    # Load and check the workbook
    from openpyxl import load_workbook
    buffer.seek(0)
    wb = load_workbook(buffer)
    
    print(f"\n  Sheets: {wb.sheetnames}")
    
    # Check TARA results sheet
    ws = wb["5-TARA分析结果"]
    print(f"  TARA Results sheet rows: {ws.max_row}")
    print(f"  TARA Results sheet columns: {ws.max_column}")
    
    # Verify data
    assert ws.max_row >= 4, "Should have at least header + 1 data row"
    print(f"\n  Sample data row 4: {[ws.cell(row=4, column=c).value for c in range(1, 6)]}")
    
    print("\n" + "=" * 60)
    print("✓ All tests passed!")
    print("=" * 60)
    
    return True


if __name__ == "__main__":
    main()
